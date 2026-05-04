"""
SFF-8472 Rev 12.5a  EEPROM Manager  v1.5.0
SFP/SFP+ Optical Transceiver EEPROM Management Tool

구조:
  CP2112I2C  - Silicon Labs CP2112 USB-HID → I2C 드라이버 (CMIS GUI와 동일)
  HexViewer  - Treeview 기반 hex 뷰어 위젯
  SFF8472App - 메인 GUI

탭 구성:
  연결    - CP2112 장치 선택 / 연결 / Password Unlock
  A0h     - 식별/설정 영역 (I2C 0xA0) hex + decoded
  A2h     - DDM/임계값 영역 (I2C 0xA2) hex + threshold + decoded
  DDM     - 실시간 측정값 모니터링
  Compare - 기준 파일 vs DUT EEPROM 비교
  로그    - 작업 이력 로그

I2C 주소 (8-bit):
  A0h = 0xA0  (7-bit 0x50)
  A2h = 0xA2  (7-bit 0x51)
"""
# ── 표준 라이브러리 ───────────────────────────────────────
import tkinter as tk
from tkinter import ttk, filedialog, messagebox, scrolledtext
import threading, time, os, sys, math, logging, colorsys, json, struct
import ctypes
from datetime import datetime

# ── 선택적 Excel 라이브러리 ──────────────────────────────
try:
    import openpyxl as _openpyxl
    _HAS_OPENPYXL = True
except ImportError:
    _HAS_OPENPYXL = False

try:
    import xlrd as _xlrd
    _HAS_XLRD = True
except ImportError:
    _HAS_XLRD = False

try:
    import xlwt as _xlwt
    _HAS_XLWT = True
except ImportError:
    _HAS_XLWT = False

# ── 앱 메타 ──────────────────────────────────────────────
APP_NAME    = "SFF-8472 EEPROM Manager"
APP_VERSION = "1.5.0"
APP_DATE    = "2026-04-21"

# ── crash.log ─────────────────────────────────────────────
def _setup_crash_log():
    log_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "crash.log")
    handler = logging.FileHandler(log_path, encoding="utf-8", mode="a")
    handler.setLevel(logging.ERROR)
    handler.setFormatter(logging.Formatter(
        "%(asctime)s [%(levelname)s] %(message)s", datefmt="%Y-%m-%d %H:%M:%S"))
    root_logger = logging.getLogger()
    root_logger.setLevel(logging.ERROR)
    root_logger.addHandler(handler)
    def _exc_handler(exc_type, exc_value, exc_tb):
        if issubclass(exc_type, KeyboardInterrupt):
            sys.__excepthook__(exc_type, exc_value, exc_tb); return
        logging.error("Uncaught exception", exc_info=(exc_type, exc_value, exc_tb))
        sys.__excepthook__(exc_type, exc_value, exc_tb)
    sys.excepthook = _exc_handler
    return log_path

CRASH_LOG = _setup_crash_log()

# ── SLABHIDtoSMBus.dll 로드 ───────────────────────────────
def _load_slab_dll():
    if sys.platform != "win32":
        return False, None
    dirs = [os.path.dirname(os.path.abspath(__file__)), os.getcwd()]
    for d in dirs:
        path = os.path.join(d, "SLABHIDtoSMBus.dll")
        if os.path.exists(path):
            try:
                return True, ctypes.WinDLL(path)
            except Exception as e:
                print(f"DLL load failed: {e}")
    print("SLABHIDtoSMBus.dll not found. Copy it to the script folder.")
    return False, None

HID_OK, _SLAB_DLL = _load_slab_dll()
CP2112_VID = 0x10C4
CP2112_PID = 0xEA90

# ── SFF-8472 I2C 주소 (8-bit) ─────────────────────────────
I2C_A0 = 0xA0   # A0h 주소 영역 (식별/설정)
I2C_A2 = 0xA2   # A2h 주소 영역 (DDM/임계값)

# ── 테마 유틸 ─────────────────────────────────────────────
def _h2rgb(h):
    h = h.lstrip("#")
    return tuple(int(h[i:i+2], 16)/255 for i in (0, 2, 4))
def _rgb2h(r, g, b):
    return "#{:02X}{:02X}{:02X}".format(int(r*255), int(g*255), int(b*255))
def _adj(hc, f):
    r, g, b = _h2rgb(hc)
    hh, s, v = colorsys.rgb_to_hsv(r, g, b)
    return _rgb2h(*colorsys.hsv_to_rgb(hh, s, max(0.0, min(1.0, v*f))))
# 강조색은 밝기 조정 제외
_SKIP_ADJ = {"acc", "acc_d", "acc_l", "grn", "red", "yel", "sel"}
def make_theme(base, f):
    return {k: (_adj(v, f) if isinstance(v, str) and v.startswith("#")
                and k not in _SKIP_ADJ else v)
            for k, v in base.items()}

# ── Dark / Light base themes (Steel Blue accent) ──────────
_DARK_BASE = {
    "bg0":    "#1C2230", "bg1": "#232B3A", "bg2": "#2C3548", "bg3": "#333F55",
    "bd":     "#3D4D68",
    "t1":     "#EEF3FA", "t2": "#B5C8E0", "t3": "#7A96B8", "t4": "#4A6080",
    "acc":    "#3A6090", "acc_d": "#243C60", "acc_l": "#5B7FA6",
    "grn":    "#4EC97A", "red": "#E05050", "yel": "#D8B840",
    "sel":    "#1E3A5A",
    "btn_bg": "#3A6090", "btn_fg": "#EEF3FA",
}
_LIGHT_BASE = {
    "bg0":    "#F4F5F7", "bg1": "#EAECF0", "bg2": "#DDE4EC", "bg3": "#D0D8E4",
    "bd":     "#BCC8D4",
    "t1":     "#1A2838", "t2": "#2A3C50", "t3": "#587898", "t4": "#7888A0",
    "acc":    "#3A5878", "acc_d": "#506888", "acc_l": "#6880A0",
    "grn":    "#1A7A40", "red": "#BB2020", "yel": "#8A6A00",
    "sel":    "#C0D0E0",
    "btn_bg": "#3A5878", "btn_fg": "#FFFFFF",
}

# Default: light theme
_THEME = dict(_LIGHT_BASE)

# ─────────────────────────────────────────────────────────
#  CP2112I2C  (CMIS GUI 구조와 동일)
# ─────────────────────────────────────────────────────────
class CP2112I2C:
    """CP2112-F03-GM SLABHIDtoSMBus.dll ctypes driver.
    Same DLL call sequence as CMIS GUI eeprom_writer_v3_4.
    """
    HID_SMBUS_SUCCESS           = 0x00
    HID_SMBUS_DEVICE_ALREADY_OPENED = 0x14  # returned when DLL handle not yet released
    XFER_IDLE         = 0x00
    XFER_BUSY         = 0x01
    XFER_COMPLETE     = 0x02
    XFER_ERROR        = 0x03

    def __init__(self, device_index=0):
        if not HID_OK or _SLAB_DLL is None:
            raise RuntimeError("SLABHIDtoSMBus.dll not found. Copy it to the script folder.")
        self._dll      = _SLAB_DLL
        self._handle   = ctypes.c_void_p(0)
        self._is_open  = False
        status = self._dll.HidSmbus_Open(
            ctypes.byref(self._handle),
            ctypes.c_uint32(device_index),
            ctypes.c_uint16(CP2112_VID),
            ctypes.c_uint16(CP2112_PID))
        # 0x14 = DEVICE_ALREADY_OPENED: DLL handle not fully released yet after close.
        # Retry up to 3 times with 50 ms delay.
        if status == self.HID_SMBUS_DEVICE_ALREADY_OPENED:
            for _ in range(3):
                time.sleep(0.05)
                status = self._dll.HidSmbus_Open(
                    ctypes.byref(self._handle),
                    ctypes.c_uint32(device_index),
                    ctypes.c_uint16(CP2112_VID),
                    ctypes.c_uint16(CP2112_PID))
                if status == self.HID_SMBUS_SUCCESS:
                    break
        if status != self.HID_SMBUS_SUCCESS:
            raise IOError(f"HidSmbus_Open failed (status=0x{status:02X})")
        self._is_open = True
        self._configure()

    def _configure(self):
        st = self._dll.HidSmbus_SetSmbusConfig(
            self._handle,
            ctypes.c_uint32(400000),  # 400 kHz
            ctypes.c_uint8(0x02),
            ctypes.c_int(0),          # autoReadRespond=False
            ctypes.c_uint16(1000),
            ctypes.c_uint16(1000),
            ctypes.c_int(1),          # sclLowTimeout=True
            ctypes.c_uint16(3))
        if st != self.HID_SMBUS_SUCCESS:
            raise IOError(f"HidSmbus_SetSmbusConfig failed (0x{st:02X})")

    def _check_open(self):
        if not self._is_open:
            raise IOError("CP2112 not connected. Please reconnect.")

    def _wait_complete(self, timeout_ms=2000):
        self._check_open()
        deadline = time.time() + timeout_ms / 1000.0
        ts = ctypes.c_uint8(0)
        ds = ctypes.c_uint8(0)
        nr = ctypes.c_uint16(0)
        br = ctypes.c_uint16(0)
        while time.time() < deadline:
            self._dll.HidSmbus_TransferStatusRequest(self._handle)
            st = self._dll.HidSmbus_GetTransferStatusResponse(
                self._handle,
                ctypes.byref(ts), ctypes.byref(ds),
                ctypes.byref(nr), ctypes.byref(br))
            if st != self.HID_SMBUS_SUCCESS:
                raise IOError(f"GetTransferStatusResponse failed (0x{st:02X})")
            v = ts.value
            if v == self.XFER_COMPLETE: return br.value
            if v == self.XFER_ERROR:
                raise IOError(f"I2C transfer error (S1=0x{ds.value:02X}). Check wiring/address/pullup.")
            time.sleep(0.002)
        raise IOError(f"CP2112 timeout ({timeout_ms}ms)")

    def close(self):
        self._is_open = False
        try: self._dll.HidSmbus_Close(self._handle)
        except Exception: pass

    def write_byte(self, i2c_addr_8bit, reg, value):
        self._check_open()
        buf = (ctypes.c_uint8 * 2)(reg & 0xFF, value & 0xFF)
        st = self._dll.HidSmbus_WriteRequest(
            self._handle,
            ctypes.c_uint8(i2c_addr_8bit & 0xFE),
            buf,
            ctypes.c_uint8(2))
        if st != self.HID_SMBUS_SUCCESS:
            raise IOError(f"HidSmbus_WriteRequest failed (0x{st:02X})")
        self._wait_complete()
        return True

    def read_page(self, i2c_addr_8bit, start_reg, num_bytes=128):
        """LabVIEW VI와 동일한 읽기 시퀀스:
        AddressReadRequest → _wait_complete → ForceReadResponse(61B씩) → GetReadResponse
        """
        self._check_open()
        target_addr = (ctypes.c_uint8 * 16)(start_reg & 0xFF)
        st = self._dll.HidSmbus_AddressReadRequest(
            self._handle,
            ctypes.c_uint8(i2c_addr_8bit & 0xFE),
            ctypes.c_uint16(num_bytes),
            ctypes.c_uint8(1),
            target_addr)
        if st != self.HID_SMBUS_SUCCESS:
            raise IOError(f"HidSmbus_AddressReadRequest failed (0x{st:02X})")
        self._wait_complete()
        remaining = num_bytes
        result    = []
        while remaining > 0:
            chunk = min(remaining, 61)
            st_f  = self._dll.HidSmbus_ForceReadResponse(
                self._handle, ctypes.c_uint16(chunk))
            if st_f != self.HID_SMBUS_SUCCESS:
                break
            rs  = ctypes.c_uint8(0)
            buf = (ctypes.c_uint8 * 61)()
            nr  = ctypes.c_uint8(0)
            st2 = self._dll.HidSmbus_GetReadResponse(
                self._handle,
                ctypes.byref(rs), buf,
                ctypes.c_uint8(61), ctypes.byref(nr))
            if st2 != self.HID_SMBUS_SUCCESS or nr.value == 0:
                break
            result.extend(buf[j] for j in range(nr.value))
            remaining -= nr.value
        while len(result) < num_bytes:
            result.append(0)
        return result[:num_bytes]

    def read_byte(self, i2c_addr_8bit, reg):
        result = self.read_page(i2c_addr_8bit, reg, 1)
        if not result:
            raise IOError(f"No read data reg=0x{reg:02X}")
        return result[0]

    @staticmethod
    def list_devices():
        """Return list of connected CP2112 devices (same as CMIS GUI)."""
        if not HID_OK or _SLAB_DLL is None:
            return []
        num = ctypes.c_uint32(0)
        _SLAB_DLL.HidSmbus_GetNumDevices(
            ctypes.byref(num),
            ctypes.c_uint16(CP2112_VID),
            ctypes.c_uint16(CP2112_PID))
        result = []
        for i in range(num.value):
            vid = ctypes.c_uint16(0)
            pid = ctypes.c_uint16(0)
            rel = ctypes.c_uint16(0)
            _SLAB_DLL.HidSmbus_GetAttributes(
                ctypes.c_uint32(i),
                ctypes.c_uint16(CP2112_VID), ctypes.c_uint16(CP2112_PID),
                ctypes.byref(vid), ctypes.byref(pid), ctypes.byref(rel))
            sn_buf = ctypes.create_string_buffer(260)
            _SLAB_DLL.HidSmbus_GetString(
                ctypes.c_uint32(i),
                ctypes.c_uint16(CP2112_VID), ctypes.c_uint16(CP2112_PID),
                sn_buf, ctypes.c_uint32(0x04))
            sn   = sn_buf.value.decode("ascii", "ignore").strip()
            prod = "CP2112"
            if sn: prod += f" S/N:{sn}"
            result.append({"product_string": prod, "serial_number": sn})
        return result

# ─────────────────────────────────────────────────────────
#  SFF-8472 디코드 테이블
# ─────────────────────────────────────────────────────────
IDENTIFIER_MAP = {
    0x00:"Unknown/Unspecified", 0x01:"GBIC",
    0x02:"Module/soldered (SFF-8472)", 0x03:"SFP/SFP+/SFP28",
    0x0B:"DWDM-SFP/SFP+",
}
EXT_ID_MAP     = {0x04:"2-wire interface ID only"}
CONNECTOR_MAP  = {
    0x00:"Unknown", 0x01:"SC", 0x07:"LC", 0x08:"MT-RJ",
    0x0B:"Optical Pigtail", 0x0C:"MPO 1x12", 0x0D:"MPO 2x16",
    0x22:"RJ45", 0x23:"No sep. connector",
}
ENCODING_MAP   = {
    0x00:"Unspecified", 0x01:"8B/10B", 0x02:"4B/5B",
    0x03:"NRZ", 0x04:"SONET scrambled", 0x05:"64B/66B",
    0x06:"Manchester", 0x07:"256B/257B", 0x08:"PAM4",
}
RATE_ID_MAP    = {
    0x00:"Unspecified", 0x01:"SFF-8079 (4/2/1G)",
    0x02:"SFF-8431 Rx", 0x04:"SFF-8431 Tx",
    0x06:"SFF-8431 Ind", 0x08:"FC-PI-5 Rx",
    0x0A:"FC-PI-5 Ind", 0x0C:"FC-PI-6", 0x0E:"10/8G CDR",
    0x10:"FC-PI-7",
}
SFF8472_COMP_MAP = {
    0x00:"Undefined", 0x01:"Rev 9.3", 0x02:"Rev 9.5",
    0x03:"Rev 10.2",  0x04:"Rev 10.4", 0x05:"Rev 11.0",
    0x06:"Rev 11.3",  0x07:"Rev 11.4", 0x08:"Rev 12.3",
    0x09:"Rev 12.4",  0x0A:"Rev 12.5",
}

# 비교 예외 기본값
# (label, indices, default_checked)
_A2 = 256
_CMP_EXCEPTIONS = [
    # ── A0h dynamic / unique fields ─────────────────────
    ("A0h CC_BASE           (63)",    [63],                          True),
    ("A0h CC_EXT            (95)",    [95],                          True),
    ("A0h Vendor SN      (68~83)",    list(range(68, 84)),           True),
    ("A0h Vendor Specific (96~127)",  list(range(96, 128)),          True),
    ("A0h Bytes        (128~255)",    list(range(128, 256)),         True),
    # ── A2h live measurement fields ──────────────────────
    ("A2h CC_DMI           (95)",     [_A2+95],                      True),
    ("A2h Temp          (96~97)",     [_A2+96,  _A2+97],             True),
    ("A2h Vcc           (98~99)",     [_A2+98,  _A2+99],             True),
    ("A2h TX Bias     (100~101)",     [_A2+100, _A2+101],            True),
    ("A2h TX Power    (102~103)",     [_A2+102, _A2+103],            True),
    ("A2h RX Power    (104~105)",     [_A2+104, _A2+105],            True),
    ("A2h Opt DDM     (106~109)",     list(range(_A2+106, _A2+110)), True),
    ("A2h Status/Ctrl    (110)",      [_A2+110],                     True),
    ("A2h Alarm Flags (112~113)",     [_A2+112, _A2+113],            True),
    ("A2h Warn Flags  (116~117)",     [_A2+116, _A2+117],            True),
    ("A2h Bytes       (120~255)",     list(range(_A2+120, _A2+256)), True),
]

# A0h / A2h 필드명 맵 (Compare 표시용) — 모듈 레벨 상수로 정의 (호출마다 재생성 방지)
_FIELD_A0 = {
    0:"Identifier", 1:"Ext ID", 2:"Connector",
    **{j:"Compliance" for j in range(3,11)},
    11:"Encoding", 12:"Rate Nom", 13:"Rate ID",
    14:"Len SM km", 15:"Len SM 100m",
    16:"Len OM2", 17:"Len OM1", 18:"Len OM4", 19:"Len OM3",
    **{j:"Vendor Name" for j in range(20,36)},
    36:"Compliance[36]",
    **{j:"Vendor OUI" for j in range(37,40)},
    **{j:"Vendor PN" for j in range(40,56)},
    **{j:"Vendor Rev" for j in range(56,60)},
    60:"Wavelength H", 61:"Wavelength L", 62:"FC Speed 2",
    63:"CC_BASE", 64:"Options[64]", 65:"Options[65]",
    66:"Rate Max", 67:"Rate Min",
    **{j:"Vendor SN" for j in range(68,84)},
    **{j:"Date Code" for j in range(84,92)},
    92:"Diag Mon Type", 93:"Enhanced Opt",
    94:"SFF-8472 Comp", 95:"CC_EXT",
    **{j:"Vendor Spec" for j in range(96,128)},
}
_FIELD_A2 = {
    **{j:"Threshold" for j in range(0,56)},
    **{j:"Cal/Enh Feat" for j in range(56,95)},
    95:"CC_DMI",
    96:"Temp MSB", 97:"Temp LSB",
    98:"Vcc MSB", 99:"Vcc LSB",
    100:"TX Bias MSB", 101:"TX Bias LSB",
    102:"TX Pwr MSB", 103:"TX Pwr LSB",
    104:"RX Pwr MSB", 105:"RX Pwr LSB",
    106:"Opt Temp MSB", 107:"Opt Temp LSB",
    108:"TEC Curr MSB", 109:"TEC Curr LSB",
    110:"Status/Ctrl", 111:"Reserved",
    112:"Alarm Flags[112]", 113:"Alarm Flags[113]",
    114:"TX EQ Ctrl", 115:"RX Emph Ctrl",
    116:"Warn Flags[116]", 117:"Warn Flags[117]",
    118:"Ext Ctrl[118]", 119:"Ext Ctrl[119]",
    **{j:"Vendor Spec" for j in range(120,127)},
    127:"Page Select",
    **{j:"User EEPROM" for j in range(128,248)},
    **{j:"Vendor Ctrl" for j in range(248,256)},
}

def _field_name(abs_idx):
    i = abs_idx if abs_idx < 256 else abs_idx - 256
    if abs_idx < 256:
        return _FIELD_A0.get(i, f"A0h[{i:02X}h]")
    else:
        return _FIELD_A2.get(i, f"A2h[{i:02X}h]")

def _parse_hex_cell(val):
    """Excel 셀 값(int/float/str)을 0x00~0xFF 정수로 변환. 실패 시 None."""
    if val is None:
        return None
    if isinstance(val, float):
        if not val.is_integer():
            return None
        val = int(val)
    s = str(val).strip().upper().rstrip('H')
    if not s:
        return None
    try:
        return int(s, 16)
    except ValueError:
        return None

# ─────────────────────────────────────────────────────────
#  EEPROM 데이터 모델
# ─────────────────────────────────────────────────────────
class EepromData:
    def __init__(self):
        self.a0: list = [0xFF] * 256
        self.a2: list = [0xFF] * 256
        self.valid_a0 = False
        self.valid_a2 = False

    def load_file(self, path: str):
        """Load EEPROM file.

        TXT format (whitespace-separated):
          128/256 rows, col1=A0h, col2=A2h (optional)

        Excel format (xlsx / xls):
          Row 1: headers (Add., A0h, A2h) — skipped
          Col A: address (ignored), Col B: A0h, Col C: A2h
          128 or 256 data rows. Values may be hex strings or integers.
        """
        ext = os.path.splitext(path)[1].lower()

        if ext in ('.xlsx', '.xlsm'):
            if not _HAS_OPENPYXL:
                raise RuntimeError(
                    "openpyxl이 설치되지 않았습니다.\n"
                    "pip install openpyxl 을 실행하세요.")
            wb = _openpyxl.load_workbook(path, read_only=True, data_only=True)
            ws = wb.active
            a0_raw, a2_raw, a2_col_count = [], [], 0
            first = True
            for row in ws.iter_rows(values_only=True):
                if first:          # 헤더 행 건너뜀
                    first = False
                    continue
                a0v = _parse_hex_cell(row[1] if len(row) > 1 else None)
                if a0v is None:
                    continue
                a0_raw.append(a0v)
                a2v = _parse_hex_cell(row[2] if len(row) > 2 else None)
                if a2v is not None:
                    a2_raw.append(a2v)
                    a2_col_count += 1
                else:
                    a2_raw.append(0xFF)
            wb.close()

        elif ext == '.xls':
            if not _HAS_XLRD:
                raise RuntimeError(
                    "xlrd가 설치되지 않았습니다.\n"
                    "pip install xlrd==1.2.0 을 실행하세요.")
            wb = _xlrd.open_workbook(path)
            ws = wb.sheet_by_index(0)
            a0_raw, a2_raw, a2_col_count = [], [], 0
            for r in range(1, ws.nrows):   # 헤더 행(0) 건너뜀
                a0v = _parse_hex_cell(ws.cell_value(r, 1) if ws.ncols > 1 else None)
                if a0v is None:
                    continue
                a0_raw.append(a0v)
                a2v = _parse_hex_cell(ws.cell_value(r, 2) if ws.ncols > 2 else None)
                if a2v is not None:
                    a2_raw.append(a2v)
                    a2_col_count += 1
                else:
                    a2_raw.append(0xFF)

        else:
            # ── TXT 파싱 (기존 로직) ────────────────────
            a0_raw, a2_raw = [], []
            a2_col_count = 0
            with open(path, "r") as f:
                for line in f:
                    line = line.strip()
                    if not line: continue
                    parts = line.split()
                    try:
                        a0_raw.append(int(parts[0], 16))
                    except (ValueError, IndexError):
                        continue
                    if len(parts) >= 2:
                        try:
                            a2_raw.append(int(parts[1], 16))
                            a2_col_count += 1
                        except ValueError:
                            a2_raw.append(0xFF)
                    else:
                        a2_raw.append(0xFF)

        n = len(a0_raw)
        if n not in (128, 256):
            raise ValueError(f"데이터 행이 128 또는 256이어야 합니다 (현재 {n}행)")

        # A2h is valid only if majority of rows had a second column
        has_a2 = (a2_col_count >= n // 2)

        # Pad to 256 bytes with FFh
        a0 = (a0_raw + [0xFF] * 256)[:256]
        a2 = (a2_raw + [0xFF] * 256)[:256]

        self.a0 = a0
        self.a2 = a2
        self.valid_a0 = True
        self.valid_a2 = has_a2

    def save_file(self, path: str):
        ext = os.path.splitext(path)[1].lower()

        if ext in ('.xlsx', '.xlsm'):
            if not _HAS_OPENPYXL:
                raise RuntimeError(
                    "openpyxl이 설치되지 않았습니다.\n"
                    "pip install openpyxl 을 실행하세요.")
            wb = _openpyxl.Workbook()
            ws = wb.active
            ws.append(["Add.", "A0h", "A2h"])
            for i in range(256):
                ws.append([i, f"{self.a0[i]:02X}", f"{self.a2[i]:02X}"])
            wb.save(path)

        elif ext == '.xls':
            if not _HAS_XLWT:
                raise RuntimeError(
                    "xlwt가 설치되지 않았습니다.\n"
                    "pip install xlwt 을 실행하세요.")
            wb = _xlwt.Workbook()
            ws = wb.add_sheet("EEPROM")
            for col, hdr in enumerate(["Add.", "A0h", "A2h"]):
                ws.write(0, col, hdr)
            for i in range(256):
                ws.write(i + 1, 0, i)
                ws.write(i + 1, 1, f"{self.a0[i]:02X}")
                ws.write(i + 1, 2, f"{self.a2[i]:02X}")
            wb.save(path)

        else:
            with open(path, "w") as f:
                for i in range(256):
                    f.write(f"{self.a0[i]:02X}\t{self.a2[i]:02X}\n")

    # ── 체크섬 ──────────────────────────────────────────
    def calc_cc_base(self): return sum(self.a0[0:63]) & 0xFF
    def calc_cc_ext(self):  return sum(self.a0[64:95]) & 0xFF
    def calc_cc_dmi(self):  return sum(self.a2[0:95])  & 0xFF

    def update_cc(self, which="all"):
        """Auto-update CC checksums before write. which: 'a0'|'a2'|'all'"""
        if which in ("a0","all"):
            self.a0[63] = self.calc_cc_base()
            self.a0[95] = self.calc_cc_ext()
        if which in ("a2","all"):
            self.a2[95] = self.calc_cc_dmi()

    def verify_cc(self):
        """Verify CC after read. Returns {name: (stored, calculated, ok)}."""
        cb = self.calc_cc_base()
        ce = self.calc_cc_ext()
        cd = self.calc_cc_dmi()
        return {
            "CC_BASE (A0h 63)": (self.a0[63], cb, self.a0[63] == cb),
            "CC_EXT  (A0h 95)": (self.a0[95], ce, self.a0[95] == ce),
            "CC_DMI  (A2h 95)": (self.a2[95], cd, self.a2[95] == cd),
        }

    @property
    def needs_addr_change(self):
        """True if Address Change Sequence required before A2h access (A0h[92] bit2=1)."""
        return bool(self.a0[92] & 0x04)

    @property
    def is_external_cal(self):
        """True if External Calibration is in use (A0h[92] bit4=1)."""
        return bool(self.a0[92] & 0x10)

    @property
    def is_internal_cal(self):
        """True if Internal Calibration is in use (A0h[92] bit5=1)."""
        return bool(self.a0[92] & 0x20)

    # ── A0h 디코딩 ──────────────────────────────────────
    def decode_a0(self):
        d = self.a0
        rows = []
        def _h(b):         return f"{b:02X}h"
        def _lk(m, v):     return m.get(v, f"Unknown ({_h(v)})")
        def _asc(s, e):    return bytes(d[s:e]).decode("ascii","replace").rstrip()
        def _r(a, n, v, dec=""):
            rows.append((f"{a:3d} ({_h(a)})", n, v, dec))

        _r(0,  "Identifier",          _h(d[0]),  _lk(IDENTIFIER_MAP, d[0]))
        _r(1,  "Ext Identifier",      _h(d[1]),  _lk(EXT_ID_MAP, d[1]))
        _r(2,  "Connector",           _h(d[2]),  _lk(CONNECTOR_MAP, d[2]))
        comp = " ".join(f"{d[i]:02X}" for i in range(3,11))
        _r(3,  "Compliance [3-10]",   comp,      "(Table 5-3)")
        _r(11, "Encoding",            _h(d[11]), _lk(ENCODING_MAP, d[11]))
        rn = d[12]
        rs = ">25.4 GBd (byte 66-67)" if rn==0xFF else ("Not specified" if rn==0 else f"{rn*100} MBd")
        _r(12, "Signaling Rate Nom",  _h(d[12]), rs)
        _r(13, "Rate Identifier",     _h(d[13]), _lk(RATE_ID_MAP, d[13]))
        _r(14, "Length SM (km)",      str(d[14]), f"{d[14]} km")
        _r(15, "Length SM (100m)",    str(d[15]), f"{d[15]*100} m")
        _r(16, "Length OM2 (10m)",    str(d[16]), f"{d[16]*10} m")
        _r(17, "Length OM1 (10m)",    str(d[17]), f"{d[17]*10} m")
        _r(18, "Length OM4/Cu",       str(d[18]), f"{d[18]*10} m (OM4) / {d[18]} m (Cu)")
        _r(19, "Length OM3/Cu+",      _h(d[19]),  "see Table 6-1")
        _r(20, "Vendor Name [20-35]", _asc(20,36),"(ASCII 16 bytes)")
        _r(37, "Vendor OUI [37-39]",  f"{d[37]:02X}:{d[38]:02X}:{d[39]:02X}", "IEEE OUI")
        _r(40, "Vendor PN [40-55]",   _asc(40,56),"(ASCII 16 bytes)")
        _r(56, "Vendor Rev [56-59]",  _asc(56,60),"(ASCII 4 bytes)")
        wl = (d[60]<<8)|d[61]
        _r(60, "Wavelength [60-61]",  str(wl),   f"{wl} nm")
        _r(62, "FC Speed 2",          _h(d[62]), "bit0=64GFC" if d[62]&1 else "")
        cc_base_calc = self.calc_cc_base()
        _r(63, "CC_BASE", _h(d[63]),
           "✓ OK" if d[63]==cc_base_calc else f"✗ FAIL (calc={cc_base_calc:02X}h)")
        # Options [64]
        o64 = d[64]; op = []
        if o64&0x40: op.append("PwrLvl4")
        if o64&0x20: op.append("PwrLvl3")
        if o64&0x10: op.append("Paging")
        if o64&0x08: op.append("CDR")
        if o64&0x04: op.append("Cooled")
        if o64&0x02: op.append("PwrLvl2")
        if o64&0x01: op.append("LinearRx")
        _r(64, "Options [64]", _h(o64), ", ".join(op) or "None")
        # Options [65]
        o65 = d[65]; op2 = []
        if o65&0x80: op2.append("RDT")
        if o65&0x40: op2.append("TunableTx")
        if o65&0x20: op2.append("RATE_SELECT")
        if o65&0x10: op2.append("TX_DISABLE")
        if o65&0x08: op2.append("TX_FAULT")
        if o65&0x02: op2.append("RX_LOS")
        if o65&0x01: op2.append("AddlPages")
        _r(65, "Options [65]", _h(o65), ", ".join(op2) or "None")
        _r(66, "Sig Rate Max", str(d[66]), f"+{d[66]}%" if d[12]!=0xFF else f"{d[66]*250} MBd")
        _r(67, "Sig Rate Min", str(d[67]), f"-{d[67]}%" if d[12]!=0xFF else f"±{d[67]}%")
        _r(68, "Vendor SN [68-83]",  _asc(68,84), "(ASCII 16 bytes)")
        _r(84, "Date Code [84-91]",  _asc(84,92), "(YYMMDD + lot)")
        dmt = d[92]; dp = []
        if dmt&0x40: dp.append("DDM")
        if dmt&0x20: dp.append("InternalCal")
        if dmt&0x10: dp.append("ExternalCal")
        dp.append("RxPwr=Avg" if dmt&0x08 else "RxPwr=OMA")
        if dmt&0x04: dp.append("AddrChange")
        if dmt&0x02: dp.append("RPM")
        _r(92, "Diag Mon Type", _h(dmt), ", ".join(dp))
        enh = d[93]; ep = []
        if enh&0x80: ep.append("AlmWrnFlags")
        if enh&0x40: ep.append("SoftTxDIS")
        if enh&0x20: ep.append("SoftTxFAULT")
        if enh&0x10: ep.append("SoftRxLOS")
        if enh&0x08: ep.append("SoftRATE_SEL")
        _r(93, "Enhanced Options", _h(enh), ", ".join(ep) or "None")
        _r(94, "SFF-8472 Compliance", _h(d[94]), _lk(SFF8472_COMP_MAP, d[94]))
        cc_ext_calc = self.calc_cc_ext()
        _r(95, "CC_EXT", _h(d[95]),
           "✓ OK" if d[95]==cc_ext_calc else f"✗ FAIL (calc={cc_ext_calc:02X}h)")
        return rows

    # ── A2h 임계값 디코딩 ────────────────────────────────
    def decode_a2_thresholds(self):
        """Return list of (addr_str, field, type_, raw_hex, physical_str)
        for A2h bytes 0~39 — 20 rows total.
        Uses module-level _THR_PHYS_DEFS for field definitions.
        """
        d = self.a2

        def _raw16(msb, lsb): return (d[msb] << 8) | d[lsb]

        def _temp(msb, lsb):
            r = _raw16(msb, lsb)
            return ((r - 0x10000) if r & 0x8000 else r) / 256.0

        def _volt(msb, lsb): return _raw16(msb, lsb) * 100e-6
        def _bias(msb, lsb): return _raw16(msb, lsb) * 2e-3
        def _pwr(msb, lsb):
            mw = _raw16(msb, lsb) * 0.1e-3
            return 10 * math.log10(mw) if mw > 0 else -40.0

        _phys_fn = {
            "Temperature": _temp,
            "Voltage":     _volt,
            "TX Bias":     _bias,
            "TX Power":    _pwr,
            "RX Power":    _pwr,
        }
        _fmt = {
            "Temperature": lambda v: f"{v:.3f} °C",
            "Voltage":     lambda v: f"{v:.4f} V",
            "TX Bias":     lambda v: f"{v:.3f} mA",
            "TX Power":    lambda v: f"{v:.3f} dBm",
            "RX Power":    lambda v: f"{v:.3f} dBm",
        }

        rows = []
        for msb, lsb, field, type_ in _THR_PHYS_DEFS:
            phys = _phys_fn[field](msb, lsb)
            rows.append((
                f"{msb:3d} ({msb:02X}h)",
                field, type_,
                f"{d[msb]:02X}h {d[lsb]:02X}h",
                _fmt[field](phys)
            ))
        return rows

    # ── A2h 상태/제어 필드 디코딩 (bytes 95~127) ─────────
    def decode_a2_status(self):
        """Return list of (addr_str, field_name, value_str, decoded_str)"""
        d = self.a2
        rows = []
        def _h(b):     return f"{b:02X}h"
        def _r(a, n, v, dec=""):
            rows.append((f"{a:3d} ({_h(a)})", n, v, dec))

        cc_dmi_calc = self.calc_cc_dmi()
        _r(95, "CC_DMI", _h(d[95]),
           "✓ OK" if d[95]==cc_dmi_calc else f"✗ FAIL (calc={cc_dmi_calc:02X}h)")

        def _pwr_dbm(msb, lsb):
            raw = (d[msb]<<8)|d[lsb]
            mw  = raw * 0.1e-3
            return f"{10*math.log10(mw):.2f} dBm" if mw > 0 else "-40.00 dBm"
        def _temp16(msb, lsb):
            raw = (d[msb]<<8)|d[lsb]
            if raw & 0x8000: raw -= 0x10000
            return f"{raw/256:.2f} °C"
        def _volt16(msb, lsb):
            return f"{((d[msb]<<8)|d[lsb])*100e-6:.4f} V"
        def _bias16(msb, lsb):
            return f"{((d[msb]<<8)|d[lsb])*2e-3:.3f} mA"

        _r(96,  "Temperature MSB",  _h(d[96]),  _temp16(96,97))
        _r(97,  "Temperature LSB",  _h(d[97]),  "")
        _r(98,  "Vcc MSB",          _h(d[98]),  _volt16(98,99))
        _r(99,  "Vcc LSB",          _h(d[99]),  "")
        _r(100, "TX Bias MSB",      _h(d[100]), _bias16(100,101))
        _r(101, "TX Bias LSB",      _h(d[101]), "")
        _r(102, "TX Power MSB",     _h(d[102]), _pwr_dbm(102,103))
        _r(103, "TX Power LSB",     _h(d[103]), "")
        _r(104, "RX Power MSB",     _h(d[104]), _pwr_dbm(104,105))
        _r(105, "RX Power LSB",     _h(d[105]), "")

        # Status/Control byte 110
        ctrl = d[110]
        ctrl_parts = []
        if ctrl & 0x80: ctrl_parts.append("TX_DIS_state")
        if ctrl & 0x40: ctrl_parts.append("Soft_TX_DIS")
        if ctrl & 0x20: ctrl_parts.append("RS1_state")
        if ctrl & 0x10: ctrl_parts.append("RS0_state")
        if ctrl & 0x08: ctrl_parts.append("Soft_RS0")
        if ctrl & 0x04: ctrl_parts.append("TX_FAULT")
        if ctrl & 0x02: ctrl_parts.append("RX_LOS")
        if ctrl & 0x01: ctrl_parts.append("Data_NOT_ready")
        _r(110, "Status/Control", _h(ctrl), ", ".join(ctrl_parts) or "OK")

        _r(111, "Reserved (SFF-8079)", _h(d[111]), "")

        # Alarm flags 112-113
        alm112 = d[112]
        alm_parts = []
        if alm112&0x80: alm_parts.append("Temp↑")
        if alm112&0x40: alm_parts.append("Temp↓")
        if alm112&0x20: alm_parts.append("Vcc↑")
        if alm112&0x10: alm_parts.append("Vcc↓")
        if alm112&0x08: alm_parts.append("Bias↑")
        if alm112&0x04: alm_parts.append("Bias↓")
        if alm112&0x02: alm_parts.append("TxPwr↑")
        if alm112&0x01: alm_parts.append("TxPwr↓")
        _r(112, "Alarm Flags [112]", _h(alm112), ", ".join(alm_parts) or "None")

        alm113 = d[113]
        alm2 = []
        if alm113&0x80: alm2.append("RxPwr↑")
        if alm113&0x40: alm2.append("RxPwr↓")
        _r(113, "Alarm Flags [113]", _h(alm113), ", ".join(alm2) or "None")

        # TX EQ / RX Emph 114-115
        eq = d[114]
        _r(114, "TX Input EQ Ctrl",  _h(eq),
           f"RATE_HIGH={eq>>4}dB  RATE_LOW={eq&0xF}dB")
        em = d[115]
        _r(115, "RX Output Emph Ctrl", _h(em),
           f"RATE_HIGH={em>>4}dB  RATE_LOW={em&0xF}dB")

        # Warn flags 116-117
        wn116 = d[116]
        wn_parts = []
        if wn116&0x80: wn_parts.append("Temp↑")
        if wn116&0x40: wn_parts.append("Temp↓")
        if wn116&0x20: wn_parts.append("Vcc↑")
        if wn116&0x10: wn_parts.append("Vcc↓")
        if wn116&0x08: wn_parts.append("Bias↑")
        if wn116&0x04: wn_parts.append("Bias↓")
        if wn116&0x02: wn_parts.append("TxPwr↑")
        if wn116&0x01: wn_parts.append("TxPwr↓")
        _r(116, "Warn Flags [116]",  _h(wn116), ", ".join(wn_parts) or "None")

        wn117 = d[117]
        wn2 = []
        if wn117&0x80: wn2.append("RxPwr↑")
        if wn117&0x40: wn2.append("RxPwr↓")
        _r(117, "Warn Flags [117]",  _h(wn117), ", ".join(wn2) or "None")

        # Extended ctrl 118-119
        ec118 = d[118]
        ec_parts = []
        if ec118&0x10: ec_parts.append("AdaptInputEQFail")
        if ec118&0x08: ec_parts.append("Soft_RS1")
        if ec118&0x04: ec_parts.append("PwrLvl4_En")
        if ec118&0x02: ec_parts.append("PwrLvl_OpState")
        if ec118&0x01: ec_parts.append("PwrLvl_Sel")
        _r(118, "Ext Ctrl/Status [118]", _h(ec118),
           ", ".join(ec_parts) or "None")
        ec119 = d[119]
        _r(119, "Ext Ctrl/Status [119]", _h(ec119), "")

        _r(127, "Page Select", _h(d[127]), f"Page {d[127]:02X}h")
        return rows
    def get_ddm(self):
        """Convert A2h bytes 96~111 to physical values.
        Automatically selects Internal Cal (bit5) or External Cal (bit4).
        """
        d = self.a2
        is_ext = self.is_external_cal

        def _raw16u(msb, lsb): return (d[msb]<<8)|d[lsb]
        def _raw16s(msb, lsb):
            v = _raw16u(msb, lsb)
            return v - 0x10000 if v & 0x8000 else v
        def _dbm(mw): return 10*math.log10(mw) if mw > 0 else -40.0

        if is_ext:
            def _f32(s): return struct.unpack(">f", bytes(d[s:s+4]))[0]
            def _ufp(a,b): return d[a] + d[b]/256.0
            def _sfp(a,b):
                v = (d[a]<<8)|d[b]; return v-0x10000 if v&0x8000 else v
            rx4=_f32(56); rx3=_f32(60); rx2=_f32(64)
            rx1=_f32(68); rx0=_f32(72)
            i_sl=_ufp(76,77); i_of=_sfp(78,79)
            p_sl=_ufp(80,81); p_of=_sfp(82,83)
            t_sl=_ufp(84,85); t_of=_sfp(86,87)
            v_sl=_ufp(88,89); v_of=_sfp(90,91)
            temp     = (t_sl * _raw16s(96,97) + t_of) / 256.0
            volt     = (v_sl * _raw16u(98,99) + v_of) * 100e-6
            bias     = (i_sl * _raw16u(100,101) + i_of) * 2e-3
            txpwr_mw = (p_sl * _raw16u(102,103) + p_of) * 0.1e-3
            rx_ad    = _raw16u(104,105)
            rxpwr_mw = (rx4*rx_ad**4 + rx3*rx_ad**3 + rx2*rx_ad**2
                        + rx1*rx_ad + rx0) * 0.1e-3
        else:
            temp     = _raw16s(96,97) / 256.0
            volt     = _raw16u(98,99) * 100e-6
            bias     = _raw16u(100,101) * 2e-3
            txpwr_mw = _raw16u(102,103) * 0.1e-3
            rxpwr_mw = _raw16u(104,105) * 0.1e-3

        ctrl = d[110]
        return {
            "temperature":  temp,
            "voltage":      volt,
            "tx_bias":      bias,
            "tx_power_dbm": _dbm(txpwr_mw), "tx_power_mw": txpwr_mw,
            "rx_power_dbm": _dbm(rxpwr_mw), "rx_power_mw": rxpwr_mw,
            "tx_disable":   bool(ctrl & 0x40),
            "tx_fault":     bool(ctrl & 0x04),
            "rx_los":       bool(ctrl & 0x02),
            "data_ready":   not bool(ctrl & 0x01),
            "alarm_112":    d[112], "alarm_113": d[113],
            "warn_116":     d[116], "warn_117":  d[117],
            "cal_type":     "External" if is_ext else "Internal",
        }



# ─────────────────────────────────────────────────────────
#  HexViewer 위젯
# ─────────────────────────────────────────────────────────
class HexViewer(tk.Frame):
    COLS = 16
    def __init__(self, parent, label="", on_cell_click=None, **kw):
        super().__init__(parent, bg=_THEME["bg1"], **kw)
        self._on_click = on_cell_click
        self._data     = [0xFF] * 256
        self._items    = []
        self._item_idx = {}   # iid → row_index for O(1) lookup

        if label:
            tk.Label(self, text=label, bg=_THEME["bg1"],
                     fg=_THEME["acc"], font=("Consolas",10,"bold")
                     ).pack(anchor="w", padx=4, pady=(4,0))

        cols = ["Addr"] + [f"{i:02X}" for i in range(16)] + ["ASCII"]
        self._tv = ttk.Treeview(self, columns=cols, show="headings",
                                height=16, style="HV.Treeview")
        self._tv.column("Addr",  width=52, anchor="e", stretch=False)
        for c in cols[1:17]:
            self._tv.column(c, width=28, anchor="center", stretch=False)
        self._tv.column("ASCII", width=130, anchor="w", stretch=False)
        for c in cols:
            self._tv.heading(c, text=c)

        sb = ttk.Scrollbar(self, orient="vertical", command=self._tv.yview)
        self._tv.configure(yscrollcommand=sb.set)
        self._tv.pack(side="left", fill="both", expand=True, padx=(4,0))
        sb.pack(side="left", fill="y", padx=(0,4))

        self._tv.tag_configure("even", background=_THEME["bg2"])
        self._tv.tag_configure("odd",  background=_THEME["bg3"])
        self._tv.tag_configure("diff", background="#4A1010",
                               foreground=_THEME["red"])
        self._tv.bind("<ButtonRelease-1>", self._on_click_cell)

    def get_col_widths(self) -> dict:
        """Return current column widths as {col_name: width}."""
        cols = ["Addr"] + [f"{i:02X}" for i in range(16)] + ["ASCII"]
        return {c: self._tv.column(c, "width") for c in cols}

    def set_col_widths(self, widths: dict):
        """Restore column widths from saved dict."""
        for col, w in widths.items():
            try:
                self._tv.column(col, width=int(w))
            except Exception: pass

    def set_data(self, data: list, diff_indices: set = None):
        self._data    = list(data)
        diff_indices  = diff_indices or set()
        for item in self._items:
            self._tv.delete(item)
        self._items.clear()
        self._item_idx.clear()
        for row in range(16):
            addr    = row * 16
            hexvals = []
            ascii_  = ""
            has_diff = False
            for col in range(16):
                idx = addr + col
                v   = data[idx] if idx < len(data) else 0xFF
                hexvals.append(f"{v:02X}")
                ascii_ += chr(v) if 32 <= v < 127 else "."
                if idx in diff_indices:
                    has_diff = True
            tag = "diff" if has_diff else ("even" if row%2==0 else "odd")
            iid = self._tv.insert("", "end",
                                  values=[f"{addr:3d} ({addr:02X}h)"] +
                                         hexvals + [ascii_],
                                  tags=(tag,))
            self._items.append(iid)
            self._item_idx[iid] = row

    def _on_click_cell(self, event):
        if not self._on_click: return
        item = self._tv.identify_row(event.y)
        col  = self._tv.identify_column(event.x)
        if not item or not col: return
        col_idx = int(col[1:]) - 1
        row_idx = self._item_idx.get(item, -1)
        if row_idx < 0 or col_idx < 1 or col_idx > 16: return
        abs_idx = row_idx * 16 + (col_idx - 1)
        if abs_idx < len(self._data):
            self._on_click(abs_idx, self._data[abs_idx], event.x_root, event.y_root)

# ─────────────────────────────────────────────────────────
#  헬퍼 함수
# ─────────────────────────────────────────────────────────
def _btn(parent, text, cmd, width=10, **kw):
    return tk.Button(parent, text=text, command=cmd,
                     bg=_THEME["btn_bg"], fg=_THEME["btn_fg"],
                     activebackground=_THEME["acc_l"],
                     activeforeground=_THEME["t1"],
                     font=("Consolas",10,"bold"),
                     relief="flat", padx=8, pady=4,
                     cursor="hand2", width=width, **kw)

def _init_styles():
    """Called once at startup — theme changes handled by _style_misc() thereafter."""
    s = ttk.Style()
    s.theme_use("default")
    t = _THEME
    s.configure("Dark.TNotebook",     background=t["bg0"])
    s.configure("Dark.TNotebook.Tab", background=t["bg2"],
                foreground=t["t3"],
                font=("Consolas",10,"bold"), padding=(12,4))
    s.map("Dark.TNotebook.Tab",
          background=[("selected", t["acc_d"])],
          foreground=[("selected", t["t1"])])
    for style in ("HV.Treeview", "Gen.Treeview"):
        rh = 18 if style == "HV.Treeview" else 19
        s.configure(style,
                    background=t["bg2"], foreground=t["t1"],
                    fieldbackground=t["bg2"], rowheight=rh,
                    font=("Consolas",9), borderwidth=0)
        s.configure(f"{style}.Heading",
                    background=t["acc_d"], foreground=t["t1"],
                    font=("Consolas",10,"bold"), relief="flat")
        s.map(style,
              background=[("selected", t["sel"])],
              foreground=[("selected", t["t1"])])

# Threshold field definitions: (msb, lsb, field_name, type_label)
# Used by both decode_a2_thresholds() and _THR_DEFS class attribute.
_THR_PHYS_DEFS = [
    ( 0,  1, "Temperature", "High Alarm"), ( 2,  3, "Temperature", "Low Alarm"),
    ( 4,  5, "Temperature", "High Warn"),  ( 6,  7, "Temperature", "Low Warn"),
    ( 8,  9, "Voltage",     "High Alarm"), (10, 11, "Voltage",     "Low Alarm"),
    (12, 13, "Voltage",     "High Warn"),  (14, 15, "Voltage",     "Low Warn"),
    (16, 17, "TX Bias",     "High Alarm"), (18, 19, "TX Bias",     "Low Alarm"),
    (20, 21, "TX Bias",     "High Warn"),  (22, 23, "TX Bias",     "Low Warn"),
    (24, 25, "TX Power",    "High Alarm"), (26, 27, "TX Power",    "Low Alarm"),
    (28, 29, "TX Power",    "High Warn"),  (30, 31, "TX Power",    "Low Warn"),
    (32, 33, "RX Power",    "High Alarm"), (34, 35, "RX Power",    "Low Alarm"),
    (36, 37, "RX Power",    "High Warn"),  (38, 39, "RX Power",    "Low Warn"),
]


def _phys_to_raw(field: str, value: float) -> int:
    """Convert physical value to 16-bit raw EEPROM integer for threshold bytes.
    SFF-8472 Table 9-11: thresholds always stored in internal calibration units.
    Returns uint16 (0~65535).
    """
    if field == "Temperature":
        # signed 16-bit fixed-point, 1/256 °C per LSB, range -128 ~ +127.996 °C
        if not (-128.0 <= value <= 127.996):
            raise ValueError(f"Temperature out of range: {value} (valid: -128 ~ +127.996 °C)")
        raw = int(round(value * 256.0)) & 0xFFFF
    elif field == "Voltage":
        # unsigned, 100 µV per LSB
        raw = int(round(value / 100e-6))
        raw = max(0, min(65535, raw))
    elif field == "TX Bias":
        # unsigned, 2 µA per LSB
        raw = int(round(value / 2e-3))
        raw = max(0, min(65535, raw))
    elif field in ("TX Power", "RX Power"):
        # unsigned, 0.1 µW per LSB; input in dBm
        mw  = 10 ** (value / 10.0)
        raw = int(round(mw / 0.1e-3))
        raw = max(0, min(65535, raw))
    else:
        raise ValueError(f"Unknown field: {field}")
    return raw


def _cc_update_msg(a0: bool, a2: bool, dut) -> str:
    """Return CC auto-update log message showing only relevant CC values."""
    parts = []
    if a0:
        parts += [f"CC_BASE={dut.a0[63]:02X}h", f"CC_EXT={dut.a0[95]:02X}h"]
    if a2:
        parts += [f"CC_DMI={dut.a2[95]:02X}h"]
    return "CC auto-updated: " + "  ".join(parts)


# Alarm/Warning flag bit definitions for DDM tab (module-level constant)
_ALM_DEFS = [
    (112,7,"Temp High Alarm"),   (112,6,"Temp Low Alarm"),
    (112,5,"Vcc High Alarm"),    (112,4,"Vcc Low Alarm"),
    (112,3,"TX Bias High Alarm"),(112,2,"TX Bias Low Alarm"),
    (112,1,"TX Pwr High Alarm"), (112,0,"TX Pwr Low Alarm"),
    (113,7,"RX Pwr High Alarm"), (113,6,"RX Pwr Low Alarm"),
    (116,7,"Temp High Warn"),    (116,6,"Temp Low Warn"),
    (116,5,"Vcc High Warn"),     (116,4,"Vcc Low Warn"),
    (116,3,"TX Bias High Warn"), (116,2,"TX Bias Low Warn"),
    (116,1,"TX Pwr High Warn"),  (116,0,"TX Pwr Low Warn"),
    (117,7,"RX Pwr High Warn"),  (117,6,"RX Pwr Low Warn"),
]


# ─────────────────────────────────────────────────────────
#  메인 애플리케이션
# ─────────────────────────────────────────────────────────
class SFF8472App(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title(f"{APP_NAME}  v{APP_VERSION}")
        self.geometry("1280x820")
        self.minsize(1050, 680)
        self.configure(bg=_THEME["bg0"])

        self._dut          = EepromData()
        self._ref          = EepromData()
        self._ref_path     = ""
        self._device_index = -1
        self._exc_vars: dict = {}
        self._ddm_auto     = tk.BooleanVar(value=False)
        self._ddm_stop     = threading.Event()
        self._ddm_thread   = None
        self._ddm_reading  = False
        self._ddm_fail_cnt = 0        # consecutive auto-refresh fail counter

        # 테마 (라이트 기본)
        self._dark         = False
        self._brightness   = 1.0

        # A0h / A2h 페이지 선택 (읽기·쓰기 공통)
        self._sel_a0 = tk.BooleanVar(value=True)
        self._sel_a2 = tk.BooleanVar(value=True)

        # Password Unlock
        self._pw_addr_vars = [tk.StringVar(value=v) for v in ["7A","7B","7C","7D"]]
        self._pw_val_vars  = [tk.StringVar(value="00") for _ in range(4)]

        self._a2_ver = 0   # incremented on every a2 data change; used for threshold cache

        # DDM 인터벌 (초)
        self._ddm_interval = tk.IntVar(value=2)
        self._ddm_blink    = False   # 깜박임 상태

        # 로그 버퍼 (로그 탭)
        self._log_lines: list = []

        # DDM 전력 단위 (dBm / mW 토글)
        self._pwr_unit = tk.StringVar(value="dBm")

        # 작업 중 버튼 잠금 플래그
        self._busy = False

        # 백업 폴더 (스크립트와 같은 위치의 backup 서브폴더)
        self._backup_dir = os.path.join(
            os.path.dirname(os.path.abspath(__file__)), "backup")

        _init_styles()
        self._build_ui()
        self.protocol("WM_DELETE_WINDOW", self._on_close)
        self._load_config()          # 설정 복원 (UI 빌드 후)
        self.after(200, self._refresh_ports)

    # ═══════════════════════════════════════════════════
    #  UI 빌드
    # ═══════════════════════════════════════════════════
    def _build_ui(self):
        self._build_header()
        tk.Frame(self, bg=_THEME["acc"], height=2).pack(fill="x")
        # 모듈 식별 표시 바
        self._module_bar = tk.Frame(self, bg=_THEME["bg0"], height=20)
        self._module_bar.pack(fill="x")
        self._module_bar.pack_propagate(False)
        self._module_lbl = tk.Label(
            self._module_bar,
            text="  No module — auto-identify after connection",
            bg=_THEME["bg0"], fg=_THEME["t4"],
            font=("Consolas",9), anchor="w")
        self._module_lbl.pack(side="left", padx=4)

        self._dut_status_lbl = tk.Label(
            self._module_bar,
            text="● Not Read",
            bg=_THEME["bg0"], fg=_THEME["t4"],
            font=("Consolas",9,"bold"))
        self._dut_status_lbl.pack(side="right", padx=12)
        self._build_page_bar()
        self._build_notebook()
        self._build_statusbar()

    def _build_page_bar(self):
        """Page selection checkboxes + Read/Write/Clear/File buttons."""
        self._page_bar = tk.Frame(self, bg=_THEME["bg0"], pady=5)
        self._page_bar.pack(fill="x", padx=4)

        tk.Label(self._page_bar, text="Page:", bg=_THEME["bg0"],
                 fg=_THEME["t3"], font=("Consolas",9)).pack(side="left", padx=(8,4))

        self._cb_a0 = tk.Checkbutton(
            self._page_bar, text="A0h (0xA0)", variable=self._sel_a0,
            bg=_THEME["bg0"], fg=_THEME["t1"], selectcolor=_THEME["bg2"],
            activebackground=_THEME["bg0"], font=("Consolas",9))
        self._cb_a0.pack(side="left", padx=4)

        self._cb_a2 = tk.Checkbutton(
            self._page_bar, text="A2h (0xA2)", variable=self._sel_a2,
            bg=_THEME["bg0"], fg=_THEME["t1"], selectcolor=_THEME["bg2"],
            activebackground=_THEME["bg0"], font=("Consolas",9))
        self._cb_a2.pack(side="left", padx=4)

        tk.Frame(self._page_bar, bg=_THEME["t4"], width=1).pack(
            side="left", fill="y", padx=8)

        _btn(self._page_bar, "▼ Read",  self._read_selected,  width=8
             ).pack(side="left", padx=2)
        _btn(self._page_bar, "▲ Write", self._write_selected, width=8
             ).pack(side="left", padx=2)

        tk.Frame(self._page_bar, bg=_THEME["t4"], width=1).pack(
            side="left", fill="y", padx=6)

        def _clear_btn(parent, text, cmd):
            return tk.Button(parent, text=text, command=cmd,
                             bg=_THEME["bg3"], fg=_THEME["t2"],
                             activebackground=_THEME["bd"],
                             activeforeground=_THEME["t1"],
                             font=("Consolas",9,"bold"),
                             relief="flat", padx=6, pady=4,
                             cursor="hand2", width=9)
        _clear_btn(self._page_bar, "Clear A0h",
                   self._clear_a0).pack(side="left", padx=2)
        _clear_btn(self._page_bar, "Clear A2h",
                   self._clear_a2).pack(side="left", padx=2)

        tk.Frame(self._page_bar, bg=_THEME["t4"], width=1).pack(
            side="left", fill="y", padx=6)

        _btn(self._page_bar, "Save File", self._save_file, width=9
             ).pack(side="left", padx=2)
        _btn(self._page_bar, "Open File", self._open_file, width=9
             ).pack(side="left", padx=2)

    def _build_header(self):
        hdr = tk.Frame(self, bg=_THEME["acc_d"], height=44)
        hdr.pack(fill="x")
        hdr.pack_propagate(False)
        tk.Label(hdr, text="SFF-8472", bg=_THEME["acc_d"], fg=_THEME["t1"],
                 font=("Consolas",16,"bold")).pack(side="left", padx=12)
        tk.Label(hdr, text="SFP+ EEPROM Manager", bg=_THEME["acc_d"],
                 fg=_THEME["acc_l"], font=("Consolas",13,"bold")).pack(side="left")
        # 다크/라이트 토글
        self._theme_btn = tk.Button(
            hdr, text="🌙 Dark", width=8, relief="flat", cursor="hand2",
            bg=_THEME["acc_d"], fg=_THEME["t2"],
            activebackground=_THEME["acc"], activeforeground=_THEME["t1"],
            font=("Consolas",9), command=self._toggle_theme)
        self._theme_btn.pack(side="right", padx=(0,8))
        tk.Label(hdr, text=f"Rev 12.5a  |  v{APP_VERSION}", bg=_THEME["acc_d"],
                 fg=_THEME["t3"], font=("Consolas",9)).pack(side="right", padx=12)
        tk.Label(hdr, text="Ref: SFF-8024 Rev 4.13", bg=_THEME["acc_d"],
                 fg=_THEME["t4"], font=("Consolas",9)).pack(side="right", padx=4)
        hid_fg  = _THEME["grn"]  if HID_OK else _THEME["red"]
        hid_msg = "✓ DLL loaded" if HID_OK else "⚠ SLABHIDtoSMBus.dll not found"
        tk.Label(hdr, text=hid_msg, bg=_THEME["acc_d"],
                 fg=hid_fg, font=("Consolas",9)).pack(side="right", padx=8)

    def _build_notebook(self):
        self._nb = ttk.Notebook(self, style="Dark.TNotebook")
        self._nb.pack(fill="both", expand=True, padx=4, pady=4)
        tabs = [
            ("connect", "  Connect  ", self._build_connect_tab),
            ("a0",      "  A0h  ",     self._build_a0_tab),
            ("a2",      "  A2h  ",     self._build_a2_tab),
            ("ddm",     "  DDM  ",     self._build_ddm_tab),
            ("compare", "  Compare  ", self._build_compare_tab),
            ("log",     "  Log  ",     self._build_log_tab),
        ]
        self._tab_frames = {}
        for key, label, builder in tabs:
            f = tk.Frame(self._nb, bg=_THEME["bg1"])
            self._nb.add(f, text=label)
            self._tab_frames[key] = f
            builder(f)

    def _build_statusbar(self):
        bar = tk.Frame(self, bg=_THEME["bg0"], height=22)
        bar.pack(fill="x", side="bottom")
        bar.pack_propagate(False)
        self._log_lbl = tk.Label(bar, text="Ready", bg=_THEME["bg0"],
                                  fg=_THEME["t3"], font=("Consolas",9), anchor="w")
        self._log_lbl.pack(side="left", padx=8)
        self._status_lbl = tk.Label(bar, text="● Disconnected", bg=_THEME["bg0"],
                                     fg=_THEME["red"], font=("Consolas",9,"bold"))
        self._status_lbl.pack(side="right", padx=10)

    # ── 연결 탭 ─────────────────────────────────────────
    def _build_connect_tab(self, p):
        pad = dict(padx=10, pady=6)

        # ── 장치 선택 그룹 ──────────────────────────────
        g1 = tk.LabelFrame(p, text=" CP2112-F03-GM USB-HID I2C Bridge ",
                            fg=_THEME["acc"], bg=_THEME["bg1"],
                            font=("Consolas",10), padx=8, pady=8)
        g1.pack(fill="x", padx=16, pady=(16,6))

        tk.Label(g1, text="Device:", bg=_THEME["bg1"],
                 fg=_THEME["t2"], font=("Consolas",10)
                 ).grid(row=0, column=0, sticky="w", **pad)
        self._port_var = tk.StringVar()
        self._port_cb  = ttk.Combobox(g1, textvariable=self._port_var,
                                       width=34, state="readonly")
        self._port_cb.grid(row=0, column=1, sticky="w", **pad)
        _btn(g1, "Refresh", self._refresh_ports, width=8
             ).grid(row=0, column=2, **pad)

        self._conn_btn = _btn(g1, "Connect", self._toggle_connect, width=10)
        self._conn_btn.grid(row=0, column=3, padx=(8,4), pady=6, sticky="nsew")

        self._conn_status = tk.Label(g1, text="● Disconnected",
                                      fg=_THEME["red"], bg=_THEME["bg1"],
                                      font=("Consolas",10,"bold"))
        self._conn_status.grid(row=0, column=4, sticky="w", padx=4)

        hid_fg  = _THEME["grn"] if HID_OK else _THEME["red"]
        hid_msg = ("✓ SLABHIDtoSMBus.dll loaded" if HID_OK
                   else "⚠ SLABHIDtoSMBus.dll not found — copy to script folder")
        tk.Label(g1, text=hid_msg, fg=hid_fg, bg=_THEME["bg1"],
                 font=("Consolas",9)).grid(row=1, column=0, columnspan=5,
                                            sticky="w", padx=10, pady=(0,4))

        gp = tk.LabelFrame(p, text=" Password Unlock (EEPROM Write Protection Release) ",
                            fg=_THEME["acc"], bg=_THEME["bg1"],
                            font=("Consolas",10), padx=8, pady=8)
        gp.pack(fill="x", padx=16, pady=6)
        padp = dict(padx=6, pady=4)

        tk.Label(gp, text="PW Address:", bg=_THEME["bg1"],
                 fg=_THEME["t2"], font=("Consolas",10),
                 width=12, anchor="w").grid(row=0, column=0, sticky="w", **padp)
        pw_addr_fr = tk.Frame(gp, bg=_THEME["bg1"])
        pw_addr_fr.grid(row=0, column=1, sticky="w", **padp)
        for var in self._pw_addr_vars:
            tk.Entry(pw_addr_fr, textvariable=var, width=4,
                     bg=_THEME["bg2"], fg=_THEME["t1"],
                     insertbackground=_THEME["t1"],
                     font=("Consolas",10), justify="center"
                     ).pack(side="left", padx=2)
        tk.Label(gp, text="(hex, 4 bytes — vendor defined)",
                 bg=_THEME["bg1"], fg=_THEME["t4"],
                 font=("Consolas",9)).grid(row=0, column=2, sticky="w", padx=4)

        tk.Label(gp, text="PW Value:", bg=_THEME["bg1"],
                 fg=_THEME["t2"], font=("Consolas",10),
                 width=12, anchor="w").grid(row=1, column=0, sticky="w", **padp)
        pw_val_fr = tk.Frame(gp, bg=_THEME["bg1"])
        pw_val_fr.grid(row=1, column=1, sticky="w", **padp)
        for var in self._pw_val_vars:
            tk.Entry(pw_val_fr, textvariable=var, width=4,
                     bg=_THEME["bg2"], fg=_THEME["t1"],
                     insertbackground=_THEME["t1"],
                     font=("Consolas",10), justify="center"
                     ).pack(side="left", padx=2)

        btn_pw = tk.Frame(gp, bg=_THEME["bg1"])
        btn_pw.grid(row=0, column=3, rowspan=2, padx=10)
        _btn(btn_pw, "Unlock", self._pw_unlock, width=9).pack(pady=(0,4))
        _btn(btn_pw, "Lock",   self._pw_lock,   width=9).pack()

        self._pw_status = tk.Label(gp, text="🔒 Status unknown",
                                    bg=_THEME["bg1"], fg=_THEME["t4"],
                                    font=("Consolas",10))
        self._pw_status.grid(row=2, column=0, columnspan=4,
                              sticky="w", padx=8, pady=(0,4))

        g2 = tk.LabelFrame(p, text=" Connection Info ", fg=_THEME["acc"],
                            bg=_THEME["bg1"], font=("Consolas",10),
                            padx=8, pady=6)
        g2.pack(fill="x", padx=16, pady=6)
        self._conn_info = tk.Text(g2, height=4, state="disabled",
                                   relief="flat", bg=_THEME["bg2"],
                                   fg=_THEME["t1"], font=("Consolas",9))
        self._conn_info.pack(fill="x")

        g3 = tk.LabelFrame(p, text=" Usage Guide ", fg=_THEME["acc"],
                            bg=_THEME["bg1"], font=("Consolas",10),
                            padx=8, pady=6)
        g3.pack(fill="both", expand=True, padx=16, pady=6)
        st = scrolledtext.ScrolledText(g3, height=8, relief="flat",
                                        bg=_THEME["bg2"], fg=_THEME["t2"],
                                        font=("Consolas",9))
        st.pack(fill="both", expand=True)
        for ln in [
            "SFF-8472 EEPROM Usage",
            "",
            "  1. Connect CP2112 USB-HID I2C bridge to PC",
            "  2. Click [Refresh] → select device → click [Connect]",
            "  3. Port stays open while connected (released on Disconnect)",
            "  4. If write-protected: enter PW address/value → click [Unlock]",
            "  5. Select page (A0h / A2h) then [▼ Read] / [▲ Write]",
            "",
            "  SFF-8472 I2C addresses:",
            "    A0h (0xA0) — ID / Vendor / Config",
            "    A2h (0xA2) — DDM / Thresholds / Control",
            "",
            f"  crash.log: {CRASH_LOG}",
        ]:
            st.insert("end", ln + "\n")
        st.configure(state="disabled")

    # ── A0h 탭 ──────────────────────────────────────────
    def _build_a0_tab(self, p):
        pane = tk.PanedWindow(p, orient="horizontal", bg=_THEME["bg1"],
                              sashwidth=4, sashrelief="flat")
        pane.pack(fill="both", expand=True, padx=4, pady=4)

        # 왼쪽: HexViewer
        left = tk.Frame(pane, bg=_THEME["bg1"])
        self._hex_a0 = HexViewer(left, label="A0h  (I2C 0xA0, 7-bit 0x50)  –  256 bytes",
                                  on_cell_click=self._byte_popup_a0)
        self._hex_a0.pack(fill="both", expand=True)
        pane.add(left, minsize=520)

        # 오른쪽: Decoded
        right = tk.Frame(pane, bg=_THEME["bg1"])
        tk.Label(right, text="Decoded Fields", bg=_THEME["bg1"],
                 fg=_THEME["acc"], font=("Consolas",10,"bold")
                 ).pack(anchor="w", padx=4, pady=(4,0))
        cols = ("Addr", "Field", "Value", "Decoded")
        self._dec_tv = ttk.Treeview(right, columns=cols,
                                     show="headings", height=24,
                                     style="Gen.Treeview")
        widths = {"Addr":80, "Field":160, "Value":80, "Decoded":220}
        for c in cols:
            self._dec_tv.heading(c, text=c)
            self._dec_tv.column(c, width=widths[c], anchor="w")
        sb2 = ttk.Scrollbar(right, orient="vertical", command=self._dec_tv.yview)
        self._dec_tv.configure(yscrollcommand=sb2.set)
        self._dec_tv.pack(side="left", fill="both", expand=True)
        sb2.pack(side="left", fill="y")
        self._dec_tv.bind("<Double-1>", self._on_decoded_click)
        pane.add(right, minsize=380)

    # ── A2h 탭 ──────────────────────────────────────────
    def _build_a2_tab(self, p):
        pane = tk.PanedWindow(p, orient="horizontal", bg=_THEME["bg1"],
                              sashwidth=4, sashrelief="flat")
        pane.pack(fill="both", expand=True, padx=4, pady=4)

        # 왼쪽: HexViewer
        left = tk.Frame(pane, bg=_THEME["bg1"])
        self._hex_a2 = HexViewer(left,
                                  label="A2h  (I2C 0xA2, 7-bit 0x51)  –  256 bytes",
                                  on_cell_click=self._byte_popup_a2)
        self._hex_a2.pack(fill="both", expand=True)
        pane.add(left, minsize=520)

        # 오른쪽: Threshold (상단) + Decoded (하단)
        right = tk.Frame(pane, bg=_THEME["bg1"])

        # 상단: Alarm/Warning Threshold
        tk.Label(right, text="Alarm / Warning Thresholds  (A2h 0~39)",
                 bg=_THEME["bg1"], fg=_THEME["acc"],
                 font=("Consolas",10,"bold")).pack(anchor="w", padx=4, pady=(4,0))

        thr_frame = tk.Frame(right, bg=_THEME["bg1"])
        thr_frame.pack(fill="x", pady=(0,4))
        cols_t = ("Addr", "Field", "Type", "Raw (hex)", "Physical Value")
        self._thr_tv = ttk.Treeview(thr_frame, columns=cols_t,
                                     show="headings", height=10,
                                     style="Gen.Treeview")
        thr_widths = {"Addr":75, "Field":120, "Type":90,
                      "Raw (hex)":70, "Physical Value":110}
        for c in cols_t:
            self._thr_tv.heading(c, text=c)
            self._thr_tv.column(c, width=thr_widths[c], anchor="w")
        sb_t = ttk.Scrollbar(thr_frame, orient="vertical", command=self._thr_tv.yview)
        self._thr_tv.configure(yscrollcommand=sb_t.set)
        self._thr_tv.pack(side="left", fill="x", expand=True, padx=(4,0))
        sb_t.pack(side="left", fill="y")
        self._thr_tv.bind("<Double-1>", self._on_threshold_click)

        tk.Frame(right, bg=_THEME["bd"], height=1).pack(fill="x", padx=4, pady=2)

        # 하단: Status / Control decoded
        tk.Label(right, text="Decoded Fields  (A2h 95~127)",
                 bg=_THEME["bg1"], fg=_THEME["acc"],
                 font=("Consolas",10,"bold")).pack(anchor="w", padx=4, pady=(2,0))

        dec_frame = tk.Frame(right, bg=_THEME["bg1"])
        dec_frame.pack(fill="both", expand=True)
        cols_d = ("Addr", "Field", "Value", "Decoded")
        self._dec2_tv = ttk.Treeview(dec_frame, columns=cols_d,
                                      show="headings", height=14,
                                      style="Gen.Treeview")
        widths_d = {"Addr":75, "Field":150, "Value":65, "Decoded":175}
        for c in cols_d:
            self._dec2_tv.heading(c, text=c)
            self._dec2_tv.column(c, width=widths_d[c], anchor="w")
        sb_d = ttk.Scrollbar(dec_frame, orient="vertical", command=self._dec2_tv.yview)
        self._dec2_tv.configure(yscrollcommand=sb_d.set)
        self._dec2_tv.pack(side="left", fill="both", expand=True, padx=(4,0))
        sb_d.pack(side="left", fill="y")
        pane.add(right, minsize=400)

    # ── DDM 탭 ──────────────────────────────────────────
    def _build_ddm_tab(self, p):
        top = tk.Frame(p, bg=_THEME["bg1"])
        top.pack(fill="x", padx=8, pady=4)
        tk.Label(top, text="Real-Time DDM  (A2h bytes 96~111)",
                 bg=_THEME["bg1"], fg=_THEME["acc"],
                 font=("Consolas",10,"bold")).pack(side="left")

        self._ddm_cb = tk.Checkbutton(
            top, text="Auto-refresh", variable=self._ddm_auto,
            command=self._toggle_ddm_auto,
            bg=_THEME["bg1"], fg=_THEME["t1"],
            selectcolor=_THEME["bg2"],
            activebackground=_THEME["bg1"],
            font=("Consolas",10))
        self._ddm_cb.pack(side="left", padx=(12,2))

        tk.Spinbox(top, textvariable=self._ddm_interval,
                   from_=1, to=60, width=3,
                   bg=_THEME["bg2"], fg=_THEME["t1"],
                   buttonbackground=_THEME["bg3"],
                   font=("Consolas",10)
                   ).pack(side="left", padx=2)
        tk.Label(top, text="s", bg=_THEME["bg1"], fg=_THEME["t3"],
                 font=("Consolas",9)).pack(side="left", padx=(0,8))

        self._ddm_timer_lbl = tk.Label(top, text="", bg=_THEME["bg1"],
                                        fg=_THEME["yel"],
                                        font=("Consolas",9,"bold"), width=8)
        self._ddm_timer_lbl.pack(side="left", padx=4)

        _btn(top, "Read Now", self._refresh_ddm, width=10).pack(side="left")

        tk.Label(top, text=" | Power:", bg=_THEME["bg1"],
                 fg=_THEME["t3"], font=("Consolas",9)).pack(side="left", padx=(8,2))
        for unit in ("dBm","mW"):
            tk.Radiobutton(top, text=unit, variable=self._pwr_unit, value=unit,
                           command=self._update_ddm_display,
                           bg=_THEME["bg1"], fg=_THEME["t1"],
                           selectcolor=_THEME["bg2"],
                           activebackground=_THEME["bg1"],
                           font=("Consolas",9)).pack(side="left", padx=1)

        self._cal_lbl = tk.Label(top, text="Cal: —", bg=_THEME["bg1"],
                                  fg=_THEME["t4"], font=("Consolas",9))
        self._cal_lbl.pack(side="right", padx=8)

        # 측정값 카드
        grid = tk.Frame(p, bg=_THEME["bg1"])
        grid.pack(fill="x", padx=12, pady=8)
        self._ddm_widgets = {}
        params = [
            ("temperature",  "Temperature", "°C",   70.0,  -5.0),
            ("voltage",      "Voltage",     "V",     3.6,   3.0),
            ("tx_bias",      "TX Bias",     "mA",  100.0,   2.0),
            ("tx_power_dbm", "TX Power",    "dBm",   3.0, -10.0),
            ("rx_power_dbm", "RX Power",    "dBm",   3.0, -20.0),
        ]
        for col, (key, label, unit, hi, lo) in enumerate(params):
            card = tk.Frame(grid, bg=_THEME["bg2"], padx=10, pady=8)
            card.grid(row=0, column=col, padx=6, pady=4, sticky="ew")
            grid.columnconfigure(col, weight=1)
            tk.Label(card, text=label, bg=_THEME["bg2"],
                     fg=_THEME["t3"], font=("Consolas",9)).pack()
            val_lbl = tk.Label(card, text="---", bg=_THEME["bg2"],
                               fg=_THEME["t1"], font=("Consolas",18,"bold"))
            val_lbl.pack()
            unit_lbl = tk.Label(card, text=unit, bg=_THEME["bg2"],
                     fg=_THEME["t3"], font=("Consolas",9))
            unit_lbl.pack()
            bar_frame = tk.Frame(card, bg=_THEME["bg0"], height=6, width=120)
            bar_frame.pack(pady=2)
            bar_frame.pack_propagate(False)
            bar_fill = tk.Frame(bar_frame, bg=_THEME["grn"], height=6)
            bar_fill.place(x=0, y=0, width=0, height=6)
            thr_lbl = tk.Label(card, text="", bg=_THEME["bg2"],
                               fg=_THEME["t4"], font=("Consolas",7))
            thr_lbl.pack()
            self._ddm_widgets[key] = {
                "label": val_lbl, "unit_lbl": unit_lbl,
                "bar": bar_fill, "bar_frame": bar_frame,
                "hi": hi, "lo": lo, "thr_lbl": thr_lbl,
                "base_unit": unit,    # original unit string (dBm/mW/etc)
            }

        # 플래그
        flags = tk.Frame(p, bg=_THEME["bg1"])
        flags.pack(fill="x", padx=12, pady=4)
        self._flag_labels = {}
        for name, key in [("TX_DISABLE","tx_disable"), ("TX_FAULT","tx_fault"),
                           ("RX_LOS","rx_los"), ("DATA_READY","data_ready")]:
            fr = tk.Frame(flags, bg=_THEME["bg2"], padx=8, pady=4)
            fr.pack(side="left", padx=4)
            lbl = tk.Label(fr, text=f"● {name}", bg=_THEME["bg2"],
                           fg=_THEME["t3"], font=("Consolas",10,"bold"))
            lbl.pack()
            self._flag_labels[key] = lbl

        # 알람 테이블
        tk.Label(p, text="Alarm / Warning Flags  (A2h bytes 112~117)",
                 bg=_THEME["bg1"], fg=_THEME["acc"],
                 font=("Consolas",10,"bold")).pack(anchor="w", padx=12, pady=(8,2))
        alm_cols = ("Byte", "Bit", "Flag", "Status")
        self._alm_tv = ttk.Treeview(p, columns=alm_cols,
                                     show="headings", height=8,
                                     style="Gen.Treeview")
        for c in alm_cols:
            self._alm_tv.heading(c, text=c)
        for c, w in [("Byte",70),("Bit",40),("Flag",220),("Status",80)]:
            self._alm_tv.column(c, width=w, anchor="w")
        self._alm_tv.tag_configure("set",   foreground=_THEME["red"])
        self._alm_tv.tag_configure("clear", foreground=_THEME["grn"])
        self._alm_tv.pack(fill="x", padx=12, pady=4)

    # ── 로그 탭 ─────────────────────────────────────────
    def _build_log_tab(self, p):
        bar = tk.Frame(p, bg=_THEME["bg1"])
        bar.pack(fill="x", padx=8, pady=4)
        tk.Label(bar, text="Operation Log", bg=_THEME["bg1"],
                 fg=_THEME["acc"], font=("Consolas",10,"bold")).pack(side="left")
        tk.Button(bar, text="Clear Log", command=self._clear_log,
                  bg=_THEME["bg3"], fg=_THEME["t2"],
                  activebackground=_THEME["bd"], activeforeground=_THEME["t1"],
                  font=("Consolas",9), relief="flat", padx=6, pady=3,
                  cursor="hand2").pack(side="right", padx=4)
        tk.Button(bar, text="Save Log", command=self._save_log,
                  bg=_THEME["bg3"], fg=_THEME["t2"],
                  activebackground=_THEME["bd"], activeforeground=_THEME["t1"],
                  font=("Consolas",9), relief="flat", padx=6, pady=3,
                  cursor="hand2").pack(side="right", padx=4)

        self._log_txt = scrolledtext.ScrolledText(
            p, state="disabled", relief="flat",
            bg=_THEME["bg2"], fg=_THEME["t1"],
            font=("Consolas",9), wrap="none")
        self._log_txt.pack(fill="both", expand=True, padx=8, pady=(0,8))
        # 색상 태그
        self._log_txt.tag_configure("err",  foreground=_THEME["red"])
        self._log_txt.tag_configure("ok",   foreground=_THEME["grn"])
        self._log_txt.tag_configure("warn", foreground=_THEME["yel"])
        self._log_txt.tag_configure("info", foreground=_THEME["t2"])

    # ── Compare 탭 ──────────────────────────────────────
    def _build_compare_tab(self, p):
        top = tk.Frame(p, bg=_THEME["bg1"])
        top.pack(fill="x", padx=8, pady=6)
        _btn(top, "Load Reference", self._load_reference, width=14
             ).pack(side="left", padx=2)
        self._ref_path_lbl = tk.Label(top, text="(no reference file)",
                                       bg=_THEME["bg1"], fg=_THEME["t3"],
                                       font=("Consolas",9))
        self._ref_path_lbl.pack(side="left", padx=6)
        _btn(top, "Save Result",  self._save_compare_result, width=10
             ).pack(side="right", padx=2)
        _btn(top, "Run Compare",  self._run_compare, width=14
             ).pack(side="right", padx=2)
        _btn(top, "Read All",     lambda: self._read_eeprom("all"), width=10
             ).pack(side="right", padx=2)

        exc_frame = tk.LabelFrame(p,
                                   text="Compare Exceptions  (checked = excluded from comparison)",
                                   bg=_THEME["bg1"], fg=_THEME["t3"],
                                   font=("Consolas",9), padx=6, pady=4)
        exc_frame.pack(fill="x", padx=8, pady=4)
        for i, (label, _, default) in enumerate(_CMP_EXCEPTIONS):
            var = tk.BooleanVar(value=default)
            tk.Checkbutton(exc_frame, text=label, variable=var,
                           bg=_THEME["bg1"], fg=_THEME["t1"],
                           selectcolor=_THEME["bg2"],
                           activebackground=_THEME["bg1"],
                           font=("Consolas",9)
                           ).grid(row=i//3, column=i%3, sticky="w", padx=8, pady=1)
            self._exc_vars[label] = var

        # 결과 요약
        self._cmp_summary = tk.Label(p, text="", bg=_THEME["bg1"],
                                      fg=_THEME["t1"], font=("Consolas",10,"bold"))
        self._cmp_summary.pack(anchor="w", padx=10, pady=2)

        # 결과 테이블
        r_cols = ("Addr", "Page", "Field", "Reference", "DUT", "Result")
        self._cmp_tv = ttk.Treeview(p, columns=r_cols, show="headings",
                                     style="Gen.Treeview")
        for c, w in [("Addr",80),("Page",50),("Field",180),
                     ("Reference",80),("DUT",80),("Result",70)]:
            self._cmp_tv.heading(c, text=c)
            self._cmp_tv.column(c, width=w, anchor="w")
        self._cmp_tv.tag_configure("pass", foreground=_THEME["grn"])
        self._cmp_tv.tag_configure("fail", foreground=_THEME["red"])
        self._cmp_tv.tag_configure("skip", foreground=_THEME["yel"])
        sb_c = ttk.Scrollbar(p, orient="vertical", command=self._cmp_tv.yview)
        self._cmp_tv.configure(yscrollcommand=sb_c.set)
        self._cmp_tv.pack(side="left", fill="both", expand=True,
                          padx=(8,0), pady=4)
        sb_c.pack(side="left", fill="y", pady=4)

    # ═══════════════════════════════════════════════════
    #  연결 (CMIS GUI와 동일한 패턴)
    # ═══════════════════════════════════════════════════
    def _refresh_ports(self):
        if not HID_OK:
            self._port_cb["values"] = ["SLABHIDtoSMBus.dll not found"]
            return
        devs = CP2112I2C.list_devices()
        if devs:
            items = [f"[{i}] {d.get('product_string','CP2112')}"
                     for i, d in enumerate(devs)]
            self._port_cb["values"] = items
            self._port_var.set(items[0])
        else:
            self._port_cb["values"] = ["No CP2112 device (check USB connection)"]

    def _toggle_connect(self):
        if self._conn_btn.cget("text") == "Disconnect":
            self._ddm_stop.set()
            self._ddm_auto.set(False)
            self._device_index = -1
            self._ddm_fail_cnt = 0
            self._conn_status.config(text="● Disconnected", fg=_THEME["red"])
            self._conn_btn.config(text="Connect")
            self._status_lbl.config(text="● Disconnected", fg=_THEME["red"])
            try:
                self._module_lbl.config(
                    text="  No module — auto-identify after connection",
                    fg=_THEME["t4"])
            except Exception: pass
            self._set_dut_status("not_read")
            self._log("Disconnected")
            return
        if not HID_OK:
            messagebox.showerror("DLL Not Found",
                "Copy SLABHIDtoSMBus.dll to the script folder.")
            return
        sel = self._port_var.get()
        try:
            idx = int(sel.split("]")[0].replace("[", "").strip())
        except:
            idx = 0
        self._log(f"Connecting: device[{idx}] ({sel})")
        # Test open → close immediately (port released), save index only
        try:
            test = CP2112I2C(device_index=idx)
            test.close()
            self._device_index = idx
            self._conn_status.config(text="● Connected", fg=_THEME["grn"])
            self._conn_btn.config(text="Disconnect")
            self._status_lbl.config(text=f"● Connected  [device {idx}]",
                                    fg=_THEME["grn"])
            self._log(f"✓ CP2112 device[{idx}] connected")
            self._ddm_fail_cnt = 0
            self._set_dut_status("not_read")
            self._update_conn_info()
        except Exception as e:
            logging.error("CP2112 connect failed", exc_info=True)
            self._log(f"Connection failed (device[{idx}]): {e}", error=True)
            messagebox.showerror("Connection Failed",
                str(e) + "\n\n"
                "Another program may be holding the CP2112 port.\n"
                "Close it and try again.")

    _DUT_STATUS = {
        "ok":            ("● DUT OK",            "#1A7A40"),
        "not_read":      ("● Not Read",           None),      # use t4
        "not_responding":("● Not Responding",     "#BB2020"),
        "i2c_error":     ("● I2C Error",          "#A06000"),
    }

    def _set_dut_status(self, status: str):
        """Update DUT status label. status: 'ok'|'not_read'|'not_responding'|'i2c_error'"""
        try:
            text, color = self._DUT_STATUS.get(status, ("● Unknown", None))
            fg = color if color else _THEME["t4"]
            self._dut_status_lbl.config(text=text, fg=fg)
        except Exception: pass


    def _popup_near(self, win, x_root: int, y_root: int, offset_x=20, offset_y=10):
        """Position popup near the given screen coordinates, keeping it on screen."""
        win.update_idletasks()
        pw, ph = win.winfo_reqwidth(), win.winfo_reqheight()
        sw, sh = self.winfo_screenwidth(), self.winfo_screenheight()
        x = min(x_root + offset_x, sw - pw - 10)
        y = min(y_root + offset_y, sh - ph - 40)
        x = max(0, x); y = max(0, y)
        win.geometry(f"+{x}+{y}")

    def _update_module_label(self):
        """Update module info label below header."""
        d = self._dut.a0
        if all(v == 0xFF for v in d[20:36]):
            return   # no vendor name populated yet
        vendor = bytes(d[20:36]).decode("ascii","replace").strip()
        pn     = bytes(d[40:56]).decode("ascii","replace").strip()
        sn     = bytes(d[68:84]).decode("ascii","replace").strip()
        comp   = SFF8472_COMP_MAP.get(d[94], f"0x{d[94]:02X}")
        text   = f"  {vendor}  |  {pn}  |  S/N: {sn}  |  {comp}"
        try:
            self._module_lbl.config(text=text, fg=_THEME["acc_l"])
        except Exception: pass

    def _open_conn(self) -> CP2112I2C:
        """Open port for operation. Raises IOError if no device selected."""
        if self._device_index < 0:
            raise IOError("No device selected. Please connect first.")
        return CP2112I2C(device_index=self._device_index)

    def _close_conn(self, conn):
        """Close port after operation."""
        if conn:
            try: conn.close()
            except Exception: pass

    def _update_conn_info(self):
        sel  = self._port_var.get()   # Combobox에 이미 표시된 항목 사용
        info = f"Device: {sel}\n"
        info += f"I2C A0h: 0xA0 (7-bit 0x50)\n"
        info += f"I2C A2h: 0xA2 (7-bit 0x51)\n"
        self._conn_info.configure(state="normal")
        self._conn_info.delete("1.0", "end")
        self._conn_info.insert("end", info)
        self._conn_info.configure(state="disabled")

    # ═══════════════════════════════════════════════════
    #  Password Unlock / Lock
    # ═══════════════════════════════════════════════════
    def _parse_pw_bytes(self, var_list):
        result = []
        for var in var_list:
            try:
                result.append(int(var.get().strip(), 16) & 0xFF)
            except ValueError:
                raise ValueError(f"Invalid hex value: {var.get()!r}")
        return result

    def _pw_unlock(self):
        if self._device_index < 0:
            messagebox.showwarning("Warning", "Please connect to CP2112 first.")
            return
        try:
            addrs = self._parse_pw_bytes(self._pw_addr_vars)
            vals  = self._parse_pw_bytes(self._pw_val_vars)
        except ValueError as e:
            messagebox.showerror("Input Error", str(e)); return
        conn = None
        try:
            conn = self._open_conn()
            for addr, val in zip(addrs, vals):
                conn.write_byte(I2C_A0, addr, val)
            self._pw_status.config(
                text="🔓 Unlock complete — EEPROM write enabled",
                fg=_THEME["grn"])
            self._log(f"Password Unlock done (addr={[f'{a:02X}h' for a in addrs]})")
        except Exception as e:
            self._pw_status.config(text=f"✗ Unlock failed: {e}", fg=_THEME["red"])
            self._log(f"Unlock failed: {e}", error=True)
            messagebox.showerror("Unlock Error", str(e))
        finally:
            self._close_conn(conn)

    def _pw_lock(self):
        """Lock — write 00 to all PW address bytes."""
        if self._device_index < 0:
            messagebox.showwarning("Warning", "Please connect to CP2112 first.")
            return
        try:
            addrs = self._parse_pw_bytes(self._pw_addr_vars)
        except ValueError as e:
            messagebox.showerror("Input Error", str(e)); return
        conn = None
        try:
            conn = self._open_conn()
            for addr in addrs:
                conn.write_byte(I2C_A0, addr, 0x00)
            self._pw_status.config(text="🔒 Lock complete", fg=_THEME["t3"])
            self._log("Password Lock done")
        except Exception as e:
            self._pw_status.config(text=f"✗ Lock failed: {e}", fg=_THEME["red"])
            self._log(f"Lock failed: {e}", error=True)
            messagebox.showerror("Lock Error", str(e))
        finally:
            self._close_conn(conn)
    # ═══════════════════════════════════════════════════
    def _read_selected(self):
        """Read A0h / A2h based on page selection checkboxes."""
        a0 = self._sel_a0.get()
        a2 = self._sel_a2.get()
        if not a0 and not a2:
            messagebox.showwarning("Warning", "Select at least one page (A0h / A2h).")
            return
        which = "all" if (a0 and a2) else ("a0" if a0 else "a2")
        self._read_eeprom(which)

    def _write_selected(self):
        a0 = self._sel_a0.get()
        a2 = self._sel_a2.get()
        if not a0 and not a2:
            messagebox.showwarning("Warning", "Select at least one page (A0h / A2h).")
            return
        if self._busy:
            return
        if self._device_index < 0:
            messagebox.showwarning("Warning", "Please connect to CP2112 first.")
            return
        pages = []
        if a0: pages.append("A0h")
        if a2: pages.append("A2h")
        msg = " + ".join(pages)
        if not messagebox.askyesno("Write Confirmation",
                f"Write [{msg}] to EEPROM?\n\n"
                "• Current EEPROM will be backed up before writing.\n"
                "• CC checksums will be auto-updated.\n"
                "• Read-back verification after write."):
            return
        self._set_busy(True)
        self._log(f"EEPROM write start ({msg})...")

        def _do():
            conn = None
            try:
                conn = self._open_conn()

                # ── 1. 쓰기 전 백업 ──────────────────────
                try:
                    os.makedirs(self._backup_dir, exist_ok=True)
                    ts_str = datetime.now().strftime("%y%m%d_%H%M%S")
                    bk_path = os.path.join(
                        self._backup_dir, f"backup_{ts_str}.txt")
                    bk_a0 = []
                    bk_a2 = []
                    if a0:
                        for start in range(0, 256, 128):
                            bk_a0.extend(conn.read_page(I2C_A0, start, 128))
                    else:
                        bk_a0 = self._dut.a0[:]   # use current buffer
                    if a2:
                        for start in range(0, 256, 128):
                            bk_a2.extend(conn.read_page(I2C_A2, start, 128))
                    else:
                        bk_a2 = self._dut.a2[:]   # use current buffer
                    with open(bk_path, "w") as f:
                        for i in range(256):
                            f.write(f"{bk_a0[i]:02X}\t{bk_a2[i]:02X}\n")
                    self.after(0, lambda: self._log(
                        f"Backup saved: {os.path.basename(bk_path)}", ok=True))
                except Exception as be:
                    self.after(0, lambda: self._log(
                        f"Backup failed (continuing): {be}", warn=True))

                # ── 2. CC 자동 갱신 ──────────────────────
                which = "all" if (a0 and a2) else ("a0" if a0 else "a2")
                self._dut.update_cc(which)
                self.after(0, lambda: self._log(_cc_update_msg(a0, a2, self._dut)))
                self.after(0, self._refresh_all_views)

                # ── 3. EEPROM 쓰기 ───────────────────────
                if a0:
                    self.after(0, lambda: self._log("Writing A0h (0xA0)..."))
                    for reg, val in enumerate(self._dut.a0):
                        conn.write_byte(I2C_A0, reg, val)
                    self.after(0, lambda: self._log("A0h write complete (256 bytes)"))
                if a2:
                    self.after(0, lambda: self._log("Writing A2h (0xA2)..."))
                    for reg, val in enumerate(self._dut.a2):
                        conn.write_byte(I2C_A2, reg, val)
                    self.after(0, lambda: self._log("A2h write complete (256 bytes)"))

                # ── 4. Read-back CC 검증 ─────────────────
                time.sleep(0.10)
                self.after(0, lambda: self._log("Verifying CC read-back..."))
                rb_fail = False
                if a0:
                    rb_base = conn.read_page(I2C_A0, 63, 1)[0]
                    rb_ext  = conn.read_page(I2C_A0, 95, 1)[0]
                    ok_base = (rb_base == self._dut.a0[63])
                    ok_ext  = (rb_ext  == self._dut.a0[95])
                    self.after(0, lambda ob=ok_base, v=rb_base:
                               self._log(f"  CC_BASE read-back: {v:02X}h {'✓' if ob else '✗ FAIL'}",
                                         error=not ob, ok=ob))
                    self.after(0, lambda oe=ok_ext, v=rb_ext:
                               self._log(f"  CC_EXT  read-back: {v:02X}h {'✓' if oe else '✗ FAIL'}",
                                         error=not oe, ok=oe))
                    if not ok_base or not ok_ext: rb_fail = True
                if a2:
                    rb_dmi  = conn.read_page(I2C_A2, 95, 1)[0]
                    ok_dmi  = (rb_dmi == self._dut.a2[95])
                    self.after(0, lambda od=ok_dmi, v=rb_dmi:
                               self._log(f"  CC_DMI  read-back: {v:02X}h {'✓' if od else '✗ FAIL'}",
                                         error=not od, ok=od))
                    if not ok_dmi: rb_fail = True

                if rb_fail:
                    self.after(0, lambda: self._log(
                        "⚠ Read-back mismatch — possible write error!", error=True))
                else:
                    self.after(0, lambda: self._log(
                        f"Write complete ✓ Read-back OK  ({msg})", ok=True))

            except Exception as e:
                logging.error("EEPROM write failed", exc_info=True)
                self.after(0, lambda: self._log(f"Write error: {e}", error=True))
                self.after(0, lambda: messagebox.showerror("Write Error", str(e)))
            finally:
                self._close_conn(conn)
                self.after(0, lambda: self._set_busy(False))

        threading.Thread(target=_do, daemon=True).start()

    # ═══════════════════════════════════════════════════
    #  EEPROM 읽기 (CC 검증 + Data Not Ready + Address Change)
    # ═══════════════════════════════════════════════════
    def _read_eeprom(self, which="all"):
        if self._device_index < 0:
            messagebox.showwarning("Warning", "Please connect to CP2112 first.")
            return
        if self._busy:
            return
        self._set_busy(True)
        self._log(f"EEPROM read start ({which.upper()})...")

        def _do():
            conn = None
            try:
                conn = self._open_conn()
                if which in ("a0", "all"):
                    self.after(0, lambda: self._log("Reading A0h (0xA0)..."))
                    data = []
                    for start in range(0, 256, 128):
                        chunk = conn.read_page(I2C_A0, start, 128)
                        data.extend(chunk)
                    self._dut.a0       = list(data[:256])
                    self._dut.valid_a0 = True

                if which in ("a2", "all"):
                    # Address Change Sequence: only if A0h already read and bit set
                    if self._dut.valid_a0 and self._dut.needs_addr_change:
                        self.after(0, lambda: self._log(
                            "Address Change Sequence required (A0h[92] bit2=1). "
                            "Writing 0x00 to A2h byte 127...", warn=True))
                        conn.write_byte(I2C_A0, 0x7F, 0x00)
                        time.sleep(0.02)

                    # Data Not Ready 체크 (byte 110 bit0)
                    self.after(0, lambda: self._log("Reading A2h (0xA2)..."))
                    ctrl_byte = conn.read_page(I2C_A2, 110, 1)
                    if ctrl_byte and (ctrl_byte[0] & 0x01):
                        self.after(0, lambda: self._log(
                            "⚠ Data Not Ready (byte 110 bit0=1) — DDM values may not be valid yet.", warn=True))

                    data = []
                    for start in range(0, 256, 128):
                        chunk = conn.read_page(I2C_A2, start, 128)
                        data.extend(chunk)
                    self._dut.a2       = list(data[:256])
                    self._dut.valid_a2 = True
                    self._a2_ver += 1

                self.after(0, self._refresh_all_views)

                # CC 검증
                cc_results = self._dut.verify_cc()
                cc_msgs = []
                cc_fail = False
                for name, (stored, calc, ok) in cc_results.items():
                    if which == "a0" and "A2h" in name: continue
                    if which == "a2" and "A0h" in name: continue
                    if ok:
                        cc_msgs.append(f"  ✓ {name}: {stored:02X}h OK")
                    else:
                        cc_msgs.append(f"  ✗ {name}: stored={stored:02X}h calc={calc:02X}h FAIL")
                        cc_fail = True
                for m in cc_msgs:
                    is_fail = "FAIL" in m
                    self.after(0, lambda msg=m, f=is_fail:
                               self._log(msg, error=f, ok=(not f)))
                if cc_fail:
                    self.after(0, lambda: self._log(
                        "⚠ CC mismatch — possible EEPROM data corruption", error=True))
                else:
                    self.after(0, lambda: self._log("Read complete ✓ (all CC OK)", ok=True))

                self.after(0, lambda: self._set_dut_status("ok"))
                self.after(0, self._update_module_label)

            except Exception as e:
                logging.error("EEPROM read failed", exc_info=True)
                err_str = str(e)
                status = "not_responding" if "S1=0x00" in err_str else "i2c_error"
                self.after(0, lambda s=status: self._set_dut_status(s))
                self.after(0, lambda: self._log(f"Read error: {e}", error=True))
                self.after(0, lambda: messagebox.showerror("Read Error", str(e)))
            finally:
                self._close_conn(conn)
                self.after(0, lambda: self._set_busy(False))

        threading.Thread(target=_do, daemon=True).start()

    # ═══════════════════════════════════════════════════
    #  뷰 갱신
    # ═══════════════════════════════════════════════════
    def _refresh_all_views(self):
        self._hex_a0.set_data(self._dut.a0)
        self._hex_a2.set_data(self._dut.a2)
        self._refresh_decoded()
        self._refresh_thresholds()
        self._refresh_a2_decoded()

    def _refresh_decoded(self):
        for item in self._dec_tv.get_children():
            self._dec_tv.delete(item)
        for row in self._dut.decode_a0():
            self._dec_tv.insert("", "end", values=row)

    def _refresh_thresholds(self):
        for item in self._thr_tv.get_children():
            self._thr_tv.delete(item)
        for row in self._dut.decode_a2_thresholds():
            self._thr_tv.insert("", "end", values=row)

    def _refresh_a2_decoded(self):
        for item in self._dec2_tv.get_children():
            self._dec2_tv.delete(item)
        for row in self._dut.decode_a2_status():
            self._dec2_tv.insert("", "end", values=row)

    def _clear_a0(self):
        if not messagebox.askyesno("Clear Confirm", "Reset all A0h data to FFh?"):
            return
        self._dut.a0 = [0xFF] * 256
        self._dut.valid_a0 = False
        self._hex_a0.set_data(self._dut.a0)
        self._refresh_decoded()
        self._log("A0h cleared")

    def _clear_a2(self):
        if not messagebox.askyesno("Clear Confirm", "Reset all A2h data to FFh?"):
            return
        self._dut.a2 = [0xFF] * 256
        self._dut.valid_a2 = False
        self._hex_a2.set_data(self._dut.a2)
        self._refresh_thresholds()
        self._refresh_a2_decoded()
        self._log("A2h cleared")

    # ═══════════════════════════════════════════════════
    #  DDM
    # ═══════════════════════════════════════════════════
    _DDM_MAX_FAIL = 3   # stop auto-refresh after N consecutive failures (~6s at 2s interval)

    def _refresh_ddm(self, from_auto=False):
        if self._device_index < 0:
            if not from_auto:
                self.after(0, lambda: messagebox.showwarning(
                    "Warning", "Please connect to CP2112 first."))
            return
        if self._ddm_reading:
            return
        self._ddm_reading = True
        def _do():
            conn = None
            try:
                conn  = self._open_conn()
                chunk = conn.read_page(I2C_A2, 96, 32)
                self._dut.a2[96:128] = list(chunk[:32])
                self._dut.valid_a2   = True
                self._a2_ver        += 1
                self._ddm_fail_cnt   = 0   # reset on success
                self.after(0, self._update_ddm_display)
                self.after(0, lambda: self._set_dut_status("ok"))
            except Exception as e:
                err_str = str(e)
                if from_auto:
                    self._ddm_fail_cnt += 1
                    if self._ddm_fail_cnt >= self._DDM_MAX_FAIL:
                        self.after(0, self._ddm_auto_stop)
                status = "not_responding" if "S1=0x00" in err_str else "i2c_error"
                self.after(0, lambda s=status: self._set_dut_status(s))
                self.after(0, lambda: self._log(f"DDM read error: {e}", error=True))
            finally:
                self._close_conn(conn)
                self._ddm_reading = False
        threading.Thread(target=_do, daemon=True).start()

    def _ddm_auto_stop(self):
        """Stop auto-refresh after consecutive failures and warn user."""
        self._ddm_stop.set()
        self._ddm_auto.set(False)
        try: self._ddm_cb.deselect()
        except Exception: pass
        self._log(f"DDM auto-refresh stopped — {self._DDM_MAX_FAIL} consecutive "
                  "read failures. Check module connection.", warn=True)
        messagebox.showwarning("DDM Auto-Refresh Stopped",
            f"Auto-refresh stopped after {self._DDM_MAX_FAIL} consecutive failures.\n\n"
            "Check that the module is connected and powered.")

    def _update_ddm_display(self):
        if not self._dut.valid_a2:
            return
        ddm = self._dut.get_ddm()
        use_mw = (self._pwr_unit.get() == "mW")

        # Threshold cache — invalidated by _a2_ver counter
        if not hasattr(self, "_thr_cache_ver") or self._thr_cache_ver != self._a2_ver:
            raw_thr = self._dut.decode_a2_thresholds()
            self._thr_cache     = {(r[1], r[2]): r[4] for r in raw_thr}
            self._thr_cache_ver = self._a2_ver
        thr_map = self._thr_cache

        # Alarm / warning bitmaps
        alm = {
            "temperature":  bool(ddm["alarm_112"] & 0xC0),
            "voltage":      bool(ddm["alarm_112"] & 0x30),
            "tx_bias":      bool(ddm["alarm_112"] & 0x0C),
            "tx_power_dbm": bool(ddm["alarm_112"] & 0x03),
            "rx_power_dbm": bool(ddm["alarm_113"] & 0xC0),
        }
        warn = {
            "temperature":  bool(ddm["warn_116"] & 0xC0),
            "voltage":      bool(ddm["warn_116"] & 0x30),
            "tx_bias":      bool(ddm["warn_116"] & 0x0C),
            "tx_power_dbm": bool(ddm["warn_116"] & 0x03),
            "rx_power_dbm": bool(ddm["warn_117"] & 0xC0),
        }

        # Theme-aware alarm colors (works for both light and dark)
        if self._dark:
            alm_bg, alm_fg  = "#5A1010", "#FF8080"
            warn_bg, warn_fg = "#4A3A10", _THEME["yel"]
        else:
            alm_bg, alm_fg  = "#FADADC", "#AA1010"
            warn_bg, warn_fg = "#FFF3CC", "#8A5500"

        _param_key = {
            "temperature": "Temperature", "voltage": "Voltage",
            "tx_bias":     "TX Bias",
            "tx_power_dbm":"TX Power",    "rx_power_dbm":"RX Power",
        }

        for key, wd in self._ddm_widgets.items():
            val = ddm.get(key, 0.0)

            # Value text + unit label
            if key == "tx_power_dbm":
                if use_mw:
                    wd["label"].config(text=f"{ddm['tx_power_mw']:.4f}")
                    wd["unit_lbl"].config(text="mW")
                else:
                    wd["label"].config(text=f"{val:.3f}")
                    wd["unit_lbl"].config(text="dBm")
            elif key == "rx_power_dbm":
                if use_mw:
                    wd["label"].config(text=f"{ddm['rx_power_mw']:.4f}")
                    wd["unit_lbl"].config(text="mW")
                else:
                    wd["label"].config(text=f"{val:.3f}")
                    wd["unit_lbl"].config(text="dBm")
            elif key == "temperature":
                wd["label"].config(text=f"{val:.2f}")
                wd["unit_lbl"].config(text=wd["base_unit"])
            elif key == "voltage":
                wd["label"].config(text=f"{val:.4f}")
                wd["unit_lbl"].config(text=wd["base_unit"])
            else:
                wd["label"].config(text=f"{val:.3f}")
                wd["unit_lbl"].config(text=wd["base_unit"])

            # Card background
            if alm.get(key):
                card_bg, fg_col = alm_bg, alm_fg
            elif warn.get(key):
                card_bg, fg_col = warn_bg, warn_fg
            else:
                card_bg, fg_col = _THEME["bg2"], _THEME["t1"]

            wd["label"].config(bg=card_bg, fg=fg_col)
            wd["unit_lbl"].config(bg=card_bg, fg=fg_col)
            for child in wd["label"].master.winfo_children():
                try: child.config(bg=card_bg)
                except Exception: pass

            # Progress bar
            hi, lo = wd["hi"], wd["lo"]
            pct  = max(0.0, min(1.0, (val-lo)/(hi-lo))) if hi != lo else 0.5
            bw   = max(1, int(wd["bar_frame"].winfo_width() * pct))
            color = (_THEME["red"] if alm.get(key) else
                     _THEME["yel"] if warn.get(key) else _THEME["grn"])
            wd["bar"].config(bg=color)
            wd["bar"].place(x=0, y=0, width=bw, height=6)

            # Threshold label
            param_key = _param_key.get(key, "")
            if param_key and "thr_lbl" in wd:
                hi_val = thr_map.get((param_key, "High Alarm"), "—")
                lo_val = thr_map.get((param_key, "Low Alarm"),  "—")
                wd["thr_lbl"].config(
                    text=f"ALM H:{hi_val}  L:{lo_val}",
                    bg=card_bg, fg=_THEME["t4"])

        # 플래그
        for key, (ok_when, _) in {
            "tx_disable": (False, "TX_DISABLE"),
            "tx_fault":   (False, "TX_FAULT"),
            "rx_los":     (False, "RX_LOS"),
            "data_ready": (True,  "DATA_READY"),
        }.items():
            ok = (ddm[key] == ok_when)
            self._flag_labels[key].config(
                fg=_THEME["grn"] if ok else _THEME["red"])

        # Cal 방식 표시
        try:
            self._cal_lbl.config(
                text=f"Cal: {ddm.get('cal_type','—')}",
                fg=_THEME["acc_l"] if ddm.get("cal_type")=="External"
                else _THEME["t3"])
        except Exception: pass

        # 알람 테이블
        for item in self._alm_tv.get_children():
            self._alm_tv.delete(item)
        bmap = {112:ddm["alarm_112"], 113:ddm["alarm_113"],
                116:ddm["warn_116"],  117:ddm["warn_117"]}
        for byte_addr, bit, name in _ALM_DEFS:
            s   = bool(bmap[byte_addr] & (1<<bit))
            tag = "set" if s else "clear"
            self._alm_tv.insert("","end",
                values=(f"{byte_addr} ({byte_addr:02X}h)",
                        str(bit), name, "SET ●" if s else "clear"),
                tags=(tag,))

    def _toggle_ddm_auto(self):
        if self._ddm_auto.get():
            if self._ddm_thread and self._ddm_thread.is_alive():
                return
            self._ddm_stop.clear()
            self._ddm_thread = threading.Thread(
                target=self._ddm_loop, daemon=True)
            self._ddm_thread.start()
            self._ddm_blink_tick()   # 깜박임 시작
        else:
            self._ddm_stop.set()
            self._ddm_timer_lbl.config(text="")

    def _ddm_blink_tick(self):
        """Blink timer label every 1s while auto-refresh is active."""
        if not self._ddm_auto.get():
            self._ddm_timer_lbl.config(text="")
            return
        interval = max(1, self._ddm_interval.get())
        self._ddm_blink = not self._ddm_blink
        if self._ddm_blink:
            self._ddm_timer_lbl.config(
                text=f"● {interval}s", fg=_THEME["yel"])
        else:
            self._ddm_timer_lbl.config(
                text=f"○ {interval}s", fg=_THEME["t4"])
        self.after(1000, self._ddm_blink_tick)

    def _ddm_loop(self):
        while not self._ddm_stop.is_set():
            interval = max(1, self._ddm_interval.get())
            if self._device_index >= 0 and not self._ddm_stop.is_set():
                self._refresh_ddm(from_auto=True)
            self._ddm_stop.wait(interval)

    # ═══════════════════════════════════════════════════
    #  Compare
    # ═══════════════════════════════════════════════════
    def _load_reference(self):
        path = filedialog.askopenfilename(
            title="Select Reference File",
            filetypes=[("EEPROM files","*.txt *.xlsx *.xlsm *.xls"),
                       ("Text files","*.txt"),
                       ("Excel files","*.xlsx *.xlsm *.xls"),
                       ("All files","*.*")])
        if not path: return
        try:
            self._ref.load_file(path)
            name = os.path.basename(path)
            self._ref_path = path
            self._ref_path_lbl.config(text=name)
            self._log(f"Reference loaded: {name}")
        except Exception as e:
            messagebox.showerror("File Error", str(e))
            self._log(f"Reference file error: {e}", error=True)

    def _run_compare(self):
        if not self._ref.valid_a0:
            messagebox.showwarning("Warning", "Load a reference file first.")
            return
        if not self._dut.valid_a0 and not self._dut.valid_a2:
            if not messagebox.askyesno("Warning",
                    "DUT EEPROM has not been read.\nContinue comparing with current data (default FFh)?"):
                return
        # 예외 인덱스 집합
        exc_set = set()
        for label, indices, default in _CMP_EXCEPTIONS:
            var = self._exc_vars.get(label)
            checked = var.get() if var is not None else default
            if checked:
                exc_set.update(indices)

        # 유효하지 않은 페이지는 자동 제외 — FFh 기본값과 비교해 false FAIL 방지
        if not self._ref.valid_a2:
            exc_set.update(range(256, 512))
            self._log("Reference has no A2h data — A2h excluded from comparison", warn=True)
        if not self._dut.valid_a2:
            exc_set.update(range(256, 512))
            self._log("DUT A2h not read — A2h excluded from comparison", warn=True)
        if not self._dut.valid_a0:
            exc_set.update(range(0, 256))
            self._log("DUT A0h not read — A0h excluded from comparison", warn=True)
        dut_flat = self._dut.a0 + self._dut.a2
        ref_flat = self._ref.a0 + self._ref.a2

        for item in self._cmp_tv.get_children():
            self._cmp_tv.delete(item)

        pass_cnt = fail_cnt = skip_cnt = 0
        diff_a0 = set(); diff_a2 = set()
        self._cmp_result_cache = []

        for idx in range(512):
            ref_v = ref_flat[idx]
            dut_v = dut_flat[idx]
            page  = "A0h" if idx < 256 else "A2h"
            byte  = idx   if idx < 256 else idx - 256
            addr  = f"{byte:3d} ({byte:02X}h)"
            field = _field_name(idx)

            if idx in exc_set:
                if ref_v != dut_v:
                    self._cmp_tv.insert("", "end", tags=("skip",),
                                        values=(addr, page, field,
                                                f"{ref_v:02X}h", f"{dut_v:02X}h",
                                                "SKIP"))
                    self._cmp_result_cache.append(
                        (addr, page, field, f"{ref_v:02X}h", f"{dut_v:02X}h", "SKIP"))
                    skip_cnt += 1
            elif ref_v != dut_v:
                self._cmp_tv.insert("", "end", tags=("fail",),
                                    values=(addr, page, field,
                                            f"{ref_v:02X}h", f"{dut_v:02X}h",
                                            "FAIL ✗"))
                self._cmp_result_cache.append(
                    (addr, page, field, f"{ref_v:02X}h", f"{dut_v:02X}h", "FAIL"))
                fail_cnt += 1
                if idx < 256: diff_a0.add(idx)
                else:          diff_a2.add(idx-256)
            else:
                pass_cnt += 1
                self._cmp_result_cache.append(
                    (addr, page, field, f"{ref_v:02X}h", f"{dut_v:02X}h", "PASS"))

        # Hex 뷰어 diff 하이라이트
        self._hex_a0.set_data(self._dut.a0, diff_a0)
        self._hex_a2.set_data(self._dut.a2, diff_a2)

        result = "PASS ✓" if fail_cnt == 0 else "FAIL ✗"
        color  = _THEME["grn"] if fail_cnt == 0 else _THEME["red"]
        self._cmp_summary.config(
            text=f"Result: {result}   PASS={pass_cnt}  FAIL={fail_cnt}  SKIP={skip_cnt}",
            fg=color)
        self._log(f"Compare done — {result}  (FAIL:{fail_cnt}  SKIP:{skip_cnt})")

    # ═══════════════════════════════════════════════════
    #  파일 저장
    # ═══════════════════════════════════════════════════
    def _save_file(self):
        path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel xlsx","*.xlsx"),
                       ("Excel xls","*.xls"),
                       ("Text files","*.txt"),
                       ("All files","*.*")],
            initialfile=f"eeprom_{datetime.now().strftime('%y%m%d_%H%M%S')}.xlsx")
        if not path: return
        try:
            self._dut.save_file(path)
            self._log(f"File saved: {os.path.basename(path)}", ok=True)
        except Exception as e:
            messagebox.showerror("Save Error", str(e))

    def _open_file(self):
        path = filedialog.askopenfilename(
            title="Open EEPROM File",
            filetypes=[("EEPROM files","*.txt *.xlsx *.xlsm *.xls"),
                       ("Text files","*.txt"),
                       ("Excel files","*.xlsx *.xlsm *.xls"),
                       ("All files","*.*")])
        if not path: return
        try:
            self._dut.load_file(path)
            self._refresh_all_views()
            self._log(f"File opened: {os.path.basename(path)}", ok=True)
        except Exception as e:
            messagebox.showerror("Open Error", str(e))

    def _save_compare_result(self):
        if not hasattr(self, "_cmp_result_cache") or not self._cmp_result_cache:
            messagebox.showwarning("Warning", "Run Compare first.")
            return
        path = filedialog.asksaveasfilename(
            defaultextension=".csv",
            filetypes=[("CSV","*.csv"),("Text","*.txt"),("All","*.*")],
            initialfile=f"compare_{datetime.now().strftime('%y%m%d_%H%M%S')}.csv")
        if not path: return
        try:
            with open(path, "w", encoding="utf-8-sig") as f:
                f.write("Addr,Page,Field,Reference,DUT,Result\n")
                for row in self._cmp_result_cache:
                    f.write(",".join(str(v) for v in row) + "\n")
            self._log(f"Compare result saved: {os.path.basename(path)}", ok=True)
        except Exception as e:
            messagebox.showerror("Save Error", str(e))

    # ═══════════════════════════════════════════════════
    #  A2h Threshold 편집
    # ═══════════════════════════════════════════════════
    # Reuse module-level constant (same order as decode_a2_thresholds rows)
    _THR_DEFS = _THR_PHYS_DEFS

    def _on_threshold_click(self, event):
        """Double-click on Physical Value column → edit popup."""
        item = self._thr_tv.identify_row(event.y)
        col  = self._thr_tv.identify_column(event.x)
        if not item or col != "#5":   # only Physical Value column
            return
        row_idx = self._thr_tv.index(item)
        if row_idx >= len(self._THR_DEFS):
            return
        msb, lsb, field, type_ = self._THR_DEFS[row_idx]
        vals = self._thr_tv.item(item, "values")
        if not vals or len(vals) < 5:
            return
        cur_phys = vals[4]
        self._show_threshold_popup(msb, lsb, field, type_, cur_phys,
                                   event.x_root, event.y_root)

    def _show_threshold_popup(self, msb, lsb, field, type_, cur_phys, x_root=None, y_root=None):
        """Popup to edit threshold physical value → auto-convert to raw hex."""
        # Unit hints per field
        units = {"Temperature":"°C  (range: -128 ~ +127.996)",
                 "Voltage":    "V   (range: 0 ~ 6.5535)",
                 "TX Bias":    "mA  (range: 0 ~ 131.07)",
                 "TX Power":   "dBm (range: -40 ~ +8.2)",
                 "RX Power":   "dBm (range: -40 ~ +8.2)"}
        unit_hint = units.get(field, "")

        win = tk.Toplevel(self, bg=_THEME["bg1"])
        win.transient(self)
        win.title(f"Edit Threshold — {field} {type_}")
        win.resizable(False, False)
        win.grab_set()
        if x_root and y_root: self._popup_near(win, x_root, y_root)

        for text in [
            f"Field  : {field}",
            f"Type   : {type_}",
            f"A2h    : byte {msb} ({msb:02X}h) / {lsb} ({lsb:02X}h)",
            f"Current: {cur_phys}",
        ]:
            tk.Label(win, text=text, bg=_THEME["bg1"], fg=_THEME["t1"],
                     font=("Consolas",10), anchor="w",
                     padx=16, pady=2).pack(fill="x")

        tk.Frame(win, bg=_THEME["bd"], height=1).pack(fill="x", padx=8, pady=4)

        fr = tk.Frame(win, bg=_THEME["bg1"])
        fr.pack(fill="x", padx=12, pady=6)
        tk.Label(fr, text=f"New value ({unit_hint}):",
                 bg=_THEME["bg1"], fg=_THEME["t3"],
                 font=("Consolas",9)).pack(anchor="w")
        entry = tk.Entry(fr, width=18, bg=_THEME["bg2"], fg=_THEME["t1"],
                         insertbackground=_THEME["t1"], font=("Consolas",11))
        entry.pack(fill="x", pady=4)
        # Strip unit from current value for default
        try:
            entry.insert(0, cur_phys.split()[0])
        except Exception: pass

        result_lbl = tk.Label(fr, text="Raw: —", bg=_THEME["bg1"],
                              fg=_THEME["t3"], font=("Consolas",9))
        result_lbl.pack(anchor="w")

        def _preview(*_):
            try:
                v = float(entry.get())
                raw = _phys_to_raw(field, v)
                msb_v, lsb_v = (raw >> 8) & 0xFF, raw & 0xFF
                result_lbl.config(
                    text=f"Raw: {raw} → {msb_v:02X}h {lsb_v:02X}h",
                    fg=_THEME["grn"])
            except Exception as e:
                result_lbl.config(text=f"Error: {e}", fg=_THEME["red"])

        entry.bind("<KeyRelease>", _preview)
        _preview()

        def _apply():
            try:
                v   = float(entry.get())
                raw = _phys_to_raw(field, v)
                msb_v = (raw >> 8) & 0xFF
                lsb_v = raw & 0xFF
                self._dut.a2[msb] = msb_v
                self._dut.a2[lsb] = lsb_v
                self._a2_ver += 1
                self._refresh_thresholds()
                self._hex_a2.set_data(self._dut.a2)
                self._log(f"Threshold {field} {type_}: "
                          f"{v} → A2h[{msb:02X}h]={msb_v:02X}h "
                          f"A2h[{lsb:02X}h]={lsb_v:02X}h", ok=True)
                win.destroy()
            except Exception as e:
                messagebox.showerror("Input Error", str(e), parent=win)

        btn_fr = tk.Frame(win, bg=_THEME["bg1"])
        btn_fr.pack(fill="x", padx=12, pady=(0,10))
        _btn(btn_fr, "Apply", _apply, width=8).pack(side="left", padx=4)
        tk.Button(btn_fr, text="Cancel", command=win.destroy,
                  bg=_THEME["bg3"], fg=_THEME["t2"],
                  font=("Consolas",9), relief="flat",
                  padx=6, pady=4, cursor="hand2",
                  width=8).pack(side="left", padx=4)
        entry.focus_set()
        entry.bind("<Return>", lambda _: _apply())
        entry.bind("<Escape>", lambda _: win.destroy())

    # ═══════════════════════════════════════════════════
    #  A0h Decoded Fields 편집
    # ═══════════════════════════════════════════════════
    # Fields editable via dropdown (byte addr → (map_dict, byte_count))
    _DECODED_DROPDOWNS = {
        0:  (IDENTIFIER_MAP, 1),
        2:  (CONNECTOR_MAP,  1),
        11: (ENCODING_MAP,   1),
        13: (RATE_ID_MAP,    1),
        94: (SFF8472_COMP_MAP, 1),
    }
    # ASCII multi-byte fields: addr → (start, end)
    _DECODED_ASCII = {
        20: (20, 36, "Vendor Name"),
        40: (40, 56, "Vendor PN"),
        56: (56, 60, "Vendor Rev"),
        68: (68, 84, "Vendor SN"),
        84: (84, 92, "Date Code"),
    }

    def _on_decoded_click(self, event):
        """Double-click on A0h Decoded Fields Treeview → edit popup."""
        item = self._dec_tv.identify_row(event.y)
        if not item:
            return
        values = self._dec_tv.item(item, "values")
        if not values:
            return
        # Parse byte address from first column "  0 (00h)"
        addr_str = str(values[0]).strip()
        try:
            byte_addr = int(addr_str.split()[0])
        except (ValueError, IndexError):
            return
        self._show_decoded_edit_popup(byte_addr, event.x_root, event.y_root)

    def _show_decoded_edit_popup(self, byte_addr, x_root=None, y_root=None):
        """Show appropriate edit popup for the given A0h byte address."""
        # Read-only fields
        if byte_addr in (63, 95):   # CC_BASE, CC_EXT
            messagebox.showinfo("Read-only",
                f"Byte {byte_addr} ({byte_addr:02X}h) is a checksum — "
                "auto-updated on write.")
            return
        if 3 <= byte_addr <= 10:    # Compliance [3-10] multi-byte
            messagebox.showinfo("Info",
                "Compliance bytes (3~10) are multi-byte fields.\n"
                "Use the HexViewer to edit individual bytes.")
            return

        # ASCII multi-byte field
        for start_b, (s, e, name) in self._DECODED_ASCII.items():
            if s <= byte_addr < e:
                self._show_ascii_popup(s, e, name, x_root, y_root)
                return

        # Dropdown field
        if byte_addr in self._DECODED_DROPDOWNS:
            map_dict, _ = self._DECODED_DROPDOWNS[byte_addr]
            self._show_dropdown_popup(byte_addr, map_dict, x_root, y_root)
            return

        # Single-byte hex edit
        self._show_single_byte_popup_dec(byte_addr, x_root, y_root)

    def _show_dropdown_popup(self, byte_addr, map_dict, x_root=None, y_root=None):
        """Edit a map-defined field via dropdown."""
        cur_val = self._dut.a0[byte_addr]
        win = tk.Toplevel(self, bg=_THEME["bg1"])
        win.transient(self)
        win.title(f"Edit  A0h [{byte_addr:02X}h]")
        win.resizable(False, False)
        win.grab_set()
        if x_root and y_root: self._popup_near(win, x_root, y_root)

        tk.Label(win, text=f"Byte {byte_addr} ({byte_addr:02X}h) — current: {cur_val:02X}h",
                 bg=_THEME["bg1"], fg=_THEME["t1"],
                 font=("Consolas",10), padx=16, pady=6).pack(fill="x")

        options = [f"{k:02X}h — {v}" for k, v in sorted(map_dict.items())]
        var = tk.StringVar()
        # Set current selection
        cur_label = f"{cur_val:02X}h — {map_dict.get(cur_val, 'Unknown')}"
        var.set(cur_label if cur_label in options else (options[0] if options else ""))

        cb = ttk.Combobox(win, textvariable=var, values=options,
                          state="readonly", width=32,
                          font=("Consolas",10))
        cb.pack(padx=16, pady=8)

        def _apply():
            sel = var.get()
            if not sel: return
            try:
                raw_hex = sel.split("h")[0].strip()
                new_val = int(raw_hex, 16)
            except ValueError:
                return
            self._dut.a0[byte_addr] = new_val
            self._dut.a0[63] = self._dut.calc_cc_base()
            self._dut.a0[95] = self._dut.calc_cc_ext()
            self._refresh_decoded()
            self._hex_a0.set_data(self._dut.a0)
            self._log(f"Decoded A0h[{byte_addr:02X}h] = {new_val:02X}h  "
                      f"({map_dict.get(new_val,'?')})  CC recalculated", ok=True)
            win.destroy()

        fr = tk.Frame(win, bg=_THEME["bg1"])
        fr.pack(padx=12, pady=(0,10))
        _btn(fr, "Apply", _apply, width=8).pack(side="left", padx=4)
        tk.Button(fr, text="Cancel", command=win.destroy,
                  bg=_THEME["bg3"], fg=_THEME["t2"],
                  font=("Consolas",9), relief="flat",
                  padx=6, pady=4, cursor="hand2",
                  width=8).pack(side="left", padx=4)
        cb.bind("<Return>", lambda _: _apply())

    def _show_ascii_popup(self, start, end, name, x_root=None, y_root=None):
        """Edit an ASCII multi-byte field."""
        cur = bytes(self._dut.a0[start:end]).decode("ascii","replace").rstrip()
        max_len = end - start
        win = tk.Toplevel(self, bg=_THEME["bg1"])
        win.transient(self)
        win.title(f"Edit  {name}")
        win.resizable(False, False)
        win.grab_set()
        if x_root and y_root: self._popup_near(win, x_root, y_root)

        tk.Label(win, text=f"{name}  (A0h bytes {start}~{end-1},  max {max_len} chars)",
                 bg=_THEME["bg1"], fg=_THEME["t1"],
                 font=("Consolas",10), padx=16, pady=6).pack(fill="x")

        entry = tk.Entry(win, width=max_len+2, bg=_THEME["bg2"], fg=_THEME["t1"],
                         insertbackground=_THEME["t1"], font=("Consolas",11))
        entry.insert(0, cur)
        entry.pack(padx=16, pady=6)

        len_lbl = tk.Label(win, text=f"0/{max_len}", bg=_THEME["bg1"],
                           fg=_THEME["t4"], font=("Consolas",9))
        len_lbl.pack(anchor="e", padx=20)

        def _upd(*_):
            n = len(entry.get())
            color = _THEME["red"] if n > max_len else _THEME["t4"]
            len_lbl.config(text=f"{n}/{max_len}", fg=color)
        entry.bind("<KeyRelease>", _upd)
        _upd()

        def _apply():
            text = entry.get()
            if len(text) > max_len:
                messagebox.showerror("Error",
                    f"Max {max_len} characters.", parent=win)
                return
            # Block non-ASCII characters
            try:
                text.encode("ascii")
            except UnicodeEncodeError:
                messagebox.showerror("Error",
                    "ASCII characters only (0x20~0x7E).\n"
                    "Remove non-ASCII characters and try again.", parent=win)
                return
            encoded = text.encode("ascii")
            padded  = encoded + b' ' * (max_len - len(encoded))
            for i, b in enumerate(padded):
                self._dut.a0[start + i] = b
            self._dut.a0[63] = self._dut.calc_cc_base()
            self._dut.a0[95] = self._dut.calc_cc_ext()
            self._refresh_decoded()
            self._hex_a0.set_data(self._dut.a0)
            self._log(f"Decoded {name} = '{text}'  CC recalculated", ok=True)
            win.destroy()

        fr = tk.Frame(win, bg=_THEME["bg1"])
        fr.pack(padx=12, pady=(0,10))
        _btn(fr, "Apply", _apply, width=8).pack(side="left", padx=4)
        tk.Button(fr, text="Cancel", command=win.destroy,
                  bg=_THEME["bg3"], fg=_THEME["t2"],
                  font=("Consolas",9), relief="flat",
                  padx=6, pady=4, cursor="hand2",
                  width=8).pack(side="left", padx=4)
        entry.focus_set()
        entry.bind("<Return>", lambda _: _apply())
        entry.bind("<Escape>", lambda _: win.destroy())

    def _show_single_byte_popup_dec(self, byte_addr, x_root=None, y_root=None):
        """Edit a single A0h byte via hex input."""
        cur_val = self._dut.a0[byte_addr]
        win = tk.Toplevel(self, bg=_THEME["bg1"])
        win.transient(self)
        win.title(f"Edit  A0h [{byte_addr:02X}h]")
        win.resizable(False, False)
        win.grab_set()
        if x_root and y_root: self._popup_near(win, x_root, y_root)

        for text in [
            f"Byte   : {byte_addr} ({byte_addr:02X}h)",
            f"Field  : {_field_name(byte_addr)}",
            f"Current: {cur_val:02X}h  ({cur_val})",
        ]:
            tk.Label(win, text=text, bg=_THEME["bg1"], fg=_THEME["t1"],
                     font=("Consolas",10), anchor="w",
                     padx=16, pady=2).pack(fill="x")

        fr = tk.Frame(win, bg=_THEME["bg1"])
        fr.pack(padx=12, pady=8)
        tk.Label(fr, text="New value (hex):", bg=_THEME["bg1"],
                 fg=_THEME["t3"], font=("Consolas",10)).pack(side="left")
        entry = tk.Entry(fr, width=6, bg=_THEME["bg2"], fg=_THEME["t1"],
                         insertbackground=_THEME["t1"], font=("Consolas",11))
        entry.insert(0, f"{cur_val:02X}")
        entry.pack(side="left", padx=6)

        def _apply():
            raw = entry.get().strip()
            try:
                new_val = int(raw, 16)
            except ValueError:
                messagebox.showerror("Error",
                    "Enter a valid hex value (00~FF).", parent=win)
                return
            if not (0 <= new_val <= 0xFF):
                messagebox.showerror("Error",
                    f"Out of range: {new_val}", parent=win)
                return
            self._dut.a0[byte_addr] = new_val
            self._dut.a0[63] = self._dut.calc_cc_base()
            self._dut.a0[95] = self._dut.calc_cc_ext()
            self._refresh_decoded()
            self._hex_a0.set_data(self._dut.a0)
            self._log(f"Decoded A0h[{byte_addr:02X}h] = {new_val:02X}h  "
                      f"CC recalculated", ok=True)
            win.destroy()

        _btn(fr, "Apply", _apply, width=8).pack(side="left", padx=4)
        tk.Button(fr, text="Cancel", command=win.destroy,
                  bg=_THEME["bg3"], fg=_THEME["t2"],
                  font=("Consolas",9), relief="flat",
                  padx=6, pady=4, cursor="hand2",
                  width=8).pack(side="left", padx=4)
        entry.focus_set()
        entry.select_range(0, "end")
        entry.bind("<Return>", lambda _: _apply())
        entry.bind("<Escape>", lambda _: win.destroy())

    def _byte_popup_a0(self, idx, val, x_root=None, y_root=None):
        self._show_byte_popup("A0h", I2C_A0, idx, val, x_root, y_root)

    def _byte_popup_a2(self, idx, val, x_root=None, y_root=None):
        self._show_byte_popup("A2h", I2C_A2, idx, val, x_root, y_root)

    def _show_byte_popup(self, page, i2c_addr, idx, val, x_root=None, y_root=None):
        win = tk.Toplevel(self, bg=_THEME["bg1"])
        win.transient(self)
        win.title(f"Byte  {page} [{idx}]")
        win.resizable(False, False)
        win.grab_set()
        if x_root and y_root: self._popup_near(win, x_root, y_root)
        abs_dec = idx if page == "A0h" else 256 + idx
        for text in [
            f"Page      : {page}",
            f"Address   : {idx:3d}  ({idx:02X}h)",
            f"Abs. Dec  : {abs_dec}",
            f"Value     : {val:3d}  ({val:02X}h)  ({val:08b}b)",
            f"Field     : {_field_name(abs_dec)}",
        ]:
            tk.Label(win, text=text, bg=_THEME["bg1"], fg=_THEME["t1"],
                     font=("Consolas",10), anchor="w",
                     padx=16, pady=2).pack(fill="x")

        fr = tk.Frame(win, bg=_THEME["bg1"])
        fr.pack(fill="x", padx=12, pady=8)
        tk.Label(fr, text="New value (hex):", bg=_THEME["bg1"],
                 fg=_THEME["t3"], font=("Consolas",10)).pack(side="left")
        entry = tk.Entry(fr, width=5, bg=_THEME["bg2"], fg=_THEME["t1"],
                         insertbackground=_THEME["t1"],
                         font=("Consolas",10))
        entry.insert(0, f"{val:02X}")
        entry.pack(side="left", padx=4)
        _btn(fr, "Write", lambda: self._write_byte(i2c_addr, idx, page, entry, win),
             width=6).pack(side="left")

    def _write_byte(self, i2c_addr, reg, page, entry_widget, win):
        if self._device_index < 0:
            messagebox.showwarning("Warning", "Not connected", parent=win)
            return
        # 입력 검증
        raw = entry_widget.get().strip()
        if not raw:
            messagebox.showwarning("Input Error", "Enter a value.", parent=win)
            return
        try:
            val = int(raw, 16)
        except ValueError:
            messagebox.showerror("Input Error",
                f"Enter a valid hex value.\nExample: 00 ~ FF", parent=win)
            return
        if not (0 <= val <= 0xFF):
            messagebox.showerror("Input Error",
                f"Value out of range: {val} (00h ~ FFh required)", parent=win)
            return
        val = val & 0xFF
        conn = None
        try:
            conn = self._open_conn()
            conn.write_byte(i2c_addr, reg, val)
            if page == "A0h":
                self._dut.a0[reg] = val
                self._dut.a0[63] = self._dut.calc_cc_base()
                self._dut.a0[95] = self._dut.calc_cc_ext()
            else:
                self._dut.a2[reg] = val
                self._dut.a2[95] = self._dut.calc_cc_dmi()
            self._refresh_all_views()
            self._log(f"Write {page} [{reg:02X}h] = {val:02X}h  CC recalculated", ok=True)
            win.destroy()
        except Exception as e:
            messagebox.showerror("Write Error", str(e), parent=win)
        finally:
            self._close_conn(conn)

    # ═══════════════════════════════════════════════════
    #  로그
    # ═══════════════════════════════════════════════════
    def _log(self, msg: str, error: bool = False, ok: bool = False,
             warn: bool = False):
        ts  = datetime.now().strftime("%H:%M:%S")
        txt = f"[{ts}] {msg}"
        # 상태바
        fg = (_THEME["red"] if error else
              _THEME["grn"] if ok else
              _THEME["yel"] if warn else _THEME["t3"])
        self._log_lbl.config(text=txt, fg=fg)
        if error: logging.error(msg)
        # 로그 탭
        self._log_lines.append(txt)
        try:
            tag = ("err" if error else "ok" if ok else
                   "warn" if warn else "info")
            self._log_txt.configure(state="normal")
            self._log_txt.insert("end", txt + "\n", tag)
            self._log_txt.see("end")
            self._log_txt.configure(state="disabled")
        except Exception: pass

    def _clear_log(self):
        self._log_lines.clear()
        try:
            self._log_txt.configure(state="normal")
            self._log_txt.delete("1.0","end")
            self._log_txt.configure(state="disabled")
        except Exception: pass

    def _save_log(self):
        path = filedialog.asksaveasfilename(
            defaultextension=".txt",
            filetypes=[("Text","*.txt"),("All","*.*")],
            initialfile=f"log_{datetime.now().strftime('%y%m%d_%H%M%S')}.txt")
        if not path: return
        try:
            with open(path,"w",encoding="utf-8") as f:
                f.write("\n".join(self._log_lines))
            self._log(f"Log saved: {os.path.basename(path)}", ok=True)
        except Exception as e:
            messagebox.showerror("Save Error", str(e))

    # ── 버튼 잠금 ────────────────────────────────────────
    def _set_busy(self, busy: bool):
        """Disable page bar buttons during read/write operation."""
        self._busy = busy
        state = "disabled" if busy else "normal"
        try:
            for attr in ("_conn_btn",):
                getattr(self, attr).config(state=state)
        except Exception: pass
        # page_bar 버튼들
        try:
            for child in self._page_bar.winfo_children():
                cls = child.winfo_class()
                if cls == "Button":
                    child.config(state=state)
        except Exception: pass

    # ═══════════════════════════════════════════════════
    #  테마
    # ═══════════════════════════════════════════════════
    def _toggle_theme(self):
        self._dark = not self._dark
        self._theme_btn.config(
            text="☀ Light" if self._dark else "🌙 Dark")
        self._apply_theme_full()

    def _apply_theme_full(self):
        global _THEME
        base = _DARK_BASE if self._dark else _LIGHT_BASE
        _THEME.update(make_theme(base, self._brightness))
        t = _THEME
        self.configure(bg=t["bg0"])
        self._walk(self, t)
        self._style_misc(t)
        # 새 위젯 직접 갱신
        try:
            self._module_bar.config(bg=t["bg0"])
            self._module_lbl.config(bg=t["bg0"])
            self._dut_status_lbl.config(bg=t["bg0"])
        except Exception: pass
        try:
            self._log_txt.config(bg=t["bg2"], fg=t["t1"])
            self._log_txt.tag_configure("err",  foreground=t["red"])
            self._log_txt.tag_configure("ok",   foreground=t["grn"])
            self._log_txt.tag_configure("warn", foreground=t["yel"])
            self._log_txt.tag_configure("info", foreground=t["t2"])
        except Exception: pass

    def _walk(self, widget, t):
        cls = widget.winfo_class()
        try:
            if cls in ("Frame", "LabelFrame"):
                # 헤더 프레임은 acc_d 색상 유지
                if hasattr(self, '_page_bar') and widget is self._page_bar:
                    widget.config(bg=t["bg0"])
                else:
                    widget.config(bg=t["bg1"])
            elif cls == "Label":
                # 헤더 영역 레이블은 acc_d 배경 유지
                try:
                    if widget.cget("bg") == _DARK_BASE["acc_d"] or \
                       widget.cget("bg") == _LIGHT_BASE["acc_d"]:
                        widget.config(bg=t["acc_d"])
                    else:
                        widget.config(bg=t["bg1"], fg=t["t1"])
                except Exception: pass
            elif cls == "Button":
                if widget is self._theme_btn:
                    widget.config(bg=t["acc_d"], fg=t["t2"],
                                   activebackground=t["acc"])
                else:
                    widget.config(bg=t["btn_bg"], fg=t["btn_fg"],
                                   activebackground=t["acc_l"],
                                   activeforeground=t["t1"])
            elif cls == "Entry":
                widget.config(bg=t["bg2"], fg=t["t1"],
                               insertbackground=t["t1"])
            elif cls == "Text":
                widget.config(bg=t["bg2"], fg=t["t1"],
                               insertbackground=t["t1"])
            elif cls == "Checkbutton":
                widget.config(bg=t["bg1"], fg=t["t1"],
                               activebackground=t["bg1"],
                               selectcolor=t["bg2"])
        except Exception: pass
        for child in widget.winfo_children():
            self._walk(child, t)

    def _style_misc(self, t):
        s = ttk.Style(self)
        s.configure("Dark.TNotebook",     background=t["bg0"])
        s.configure("Dark.TNotebook.Tab", background=t["bg2"],
                    foreground=t["t3"], font=("Consolas",10,"bold"),
                    padding=(12,4))
        s.map("Dark.TNotebook.Tab",
              background=[("selected", t["acc_d"])],
              foreground=[("selected", t["t1"])])
        for style in ("HV.Treeview", "Gen.Treeview"):
            s.configure(style,
                        background=t["bg2"], foreground=t["t1"],
                        fieldbackground=t["bg2"])
            s.configure(f"{style}.Heading",
                        background=t["acc_d"], foreground=t["t1"])
            s.map(style,
                  background=[("selected", t["sel"])],
                  foreground=[("selected", t["t1"])])
        s.configure("TCombobox",
                    fieldbackground=t["bg2"], background=t["bg2"],
                    foreground=t["t1"], selectbackground=t["sel"])
        s.configure("Vertical.TScrollbar",
                    background=t["bg1"], troughcolor=t["bg0"],
                    arrowcolor=t["t3"])

    # ═══════════════════════════════════════════════════
    #  설정 저장 / 복원  (.sff8472_config.json)
    # ═══════════════════════════════════════════════════
    def _config_path(self):
        return os.path.join(
            os.path.dirname(os.path.abspath(__file__)),
            ".sff8472_config.json")

    def _save_config(self):
        try:
            cfg = {
                "device_index": self._device_index,
                "ref_path":     self._ref_path,
                "exc_states":   {k: v.get() for k, v in self._exc_vars.items()},
                "sel_a0":       self._sel_a0.get(),
                "sel_a2":       self._sel_a2.get(),
                "dark_mode":    self._dark,
                "geometry":     self.geometry(),
                "ddm_interval": self._ddm_interval.get(),
                "pw_addr":      [v.get() for v in self._pw_addr_vars],
                "pw_val":       [v.get() for v in self._pw_val_vars],
                "hex_a0_cols":  self._hex_a0.get_col_widths(),
                "hex_a2_cols":  self._hex_a2.get_col_widths(),
            }
            with open(self._config_path(), "w", encoding="utf-8") as f:
                json.dump(cfg, f, ensure_ascii=False, indent=2)
        except Exception as e:
            logging.error(f"Config save failed: {e}")

    def _load_config(self):
        path = self._config_path()
        if not os.path.exists(path):
            return
        try:
            with open(path, "r", encoding="utf-8") as f:
                cfg = json.load(f)

            # 창 크기/위치
            geo = cfg.get("geometry")
            if geo:
                try: self.geometry(geo)
                except Exception: pass

            # 테마
            dark = cfg.get("dark_mode", False)
            if dark != self._dark:
                self._dark = dark
                self._theme_btn.config(
                    text="☀ Light" if self._dark else "🌙 Dark")
                self._apply_theme_full()

            # 페이지 선택
            self._sel_a0.set(cfg.get("sel_a0", True))
            self._sel_a2.set(cfg.get("sel_a2", True))

            # Compare 예외 체크박스
            exc = cfg.get("exc_states", {})
            for label, var in self._exc_vars.items():
                if label in exc:
                    var.set(exc[label])

            # DDM 인터벌
            self._ddm_interval.set(cfg.get("ddm_interval", 2))

            # Password 복원
            for var, val in zip(self._pw_addr_vars, cfg.get("pw_addr", [])):
                var.set(val)
            for var, val in zip(self._pw_val_vars, cfg.get("pw_val", [])):
                var.set(val)

            # 기준 파일 복원
            ref = cfg.get("ref_path", "")
            if ref and os.path.exists(ref):
                try:
                    self._ref.load_file(ref)
                    self._ref_path = ref
                    self._ref_path_lbl.config(text=os.path.basename(ref))
                    self._log(f"Reference file restored: {os.path.basename(ref)}")
                except Exception: pass

            # 장치 인덱스 (포트 새로고침 후 선택)
            di = cfg.get("device_index", -1)
            if di >= 0:
                self.after(500, lambda: self._restore_device(di))

            # HexViewer column widths
            if "hex_a0_cols" in cfg:
                self._hex_a0.set_col_widths(cfg["hex_a0_cols"])
            if "hex_a2_cols" in cfg:
                self._hex_a2.set_col_widths(cfg["hex_a2_cols"])

            self._log("Settings restored")
        except Exception as e:
            logging.error(f"Config restore failed: {e}")

    def _restore_device(self, idx):
        """Auto-select previously used device index after port refresh."""
        items = self._port_cb["values"] if hasattr(self, "_port_cb") else []
        for i, item in enumerate(items):
            try:
                if int(item.split("]")[0].replace("[","").strip()) == idx:
                    self._port_cb.current(i)
                    self._port_var.set(item)
                    return
            except: continue
    def _on_close(self):
        self._ddm_stop.set()
        self._save_config()
        self.destroy()


# ─────────────────────────────────────────────────────────
#  Entry Point
# ─────────────────────────────────────────────────────────
if __name__ == "__main__":
    app = SFF8472App()
    app.mainloop()
